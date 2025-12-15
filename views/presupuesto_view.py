from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QHeaderView, QPushButton, QLineEdit, QLabel, QMessageBox, QFileDialog,
    QInputDialog, QApplication, QDialog, QToolButton, QMenu, QSizePolicy
)
from PyQt6.QtCore import Qt, pyqtSignal
from PyQt6.QtGui import QColor, QFont, QAction
import os
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font as XLFont, PatternFill as XLFill, Alignment as XLAlign, Border as XLBorder, Side as XLSide
from controllers.analisis_unitarios_controller import AnalisisUnitariosController
import csv
import re
from models.profesional import Profesional
from models.database import SessionLocal
from models.analisis_unitario import AnalisisUnitario
from models.analisis_unitario_recurso import AnalisisUnitarioRecurso
from views.administracion_window import AdministracionWindow
from views.importar_por_texto_dialog import ImportarPorTextoDialog
from views.analisis_match_dialog import AnalisisMatchDialog

class PresupuestoView(QWidget):
    analisis_selected = pyqtSignal(str)
    analysis_edit_requested = pyqtSignal(str)
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Presupuesto")
        self.resize(1000, 600)
        self.layout = QVBoxLayout(self)
        self.chapter_counter = 0
        self.admin_cost_total = 0.0
        self.direct_cost_total = 0.0
        
        # Crear el buscador de análisis
        self.create_search_bar()
        self.create_buttons()
        self.create_table()
        
        # Crear y configurar el controlador de análisis
        self.analisis_controller = None
        
        self.setLayout(self.layout)
        self.setStyleSheet("""
            QTableWidget {
                background-color: #f9f9f9;
                gridline-color: #cccccc;
                font-size: 14px;
            }
            QHeaderView::section {
                background-color: #0078d7;
                color: white;
                padding: 4px;
                font-weight: bold;
            }
            QPushButton {
                background-color: #007ACC;
                color: white;
                border-radius: 4px;
                padding: 8px;
                min-width: 120px;
            }
            QPushButton:hover {
                background-color: #005A9E;
            }
            QToolButton {
                background-color: #007ACC;
                color: white;
                border-radius: 4px;
                padding: 8px;
                min-width: 140px;
            }
            QToolButton:hover { background-color: #005A9E; }
            QMenu {
                background-color: #007ACC;
                color: white;
                border: 1px solid #005A9E;
                padding: 0px;
                border-radius: 4px;
                font-size: 13px;
            }
            QMenu::item {
                background-color: #007ACC;
                color: white;
                padding: 8px 12px;
            }
            QMenu::item:selected {
                background-color: #005A9E;
            }
            QLineEdit {
                padding: 6px;
                border: 1px solid #cccccc;
                border-radius: 4px;
                min-width: 200px;
            }
        """)

    def create_search_bar(self):
        """Crea la barra de búsqueda de análisis unitarios."""
        search_layout = QHBoxLayout()
        
        # Campo de búsqueda por código
        self.codigo_search = QLineEdit()
        self.codigo_search.setPlaceholderText("Buscar por código")
        
        # Campo de búsqueda por descripción
        self.descripcion_search = QLineEdit()
        self.descripcion_search.setPlaceholderText("Buscar por descripción")
        
        # Botón de búsqueda
        self.search_button = QPushButton("Buscar Análisis")
        self.search_button.clicked.connect(self.show_analisis_search)
        
        # Agregar widgets al layout
        search_layout.addWidget(QLabel("Código:"))
        search_layout.addWidget(self.codigo_search)
        search_layout.addWidget(QLabel("Descripción:"))
        search_layout.addWidget(self.descripcion_search)
        search_layout.addWidget(self.search_button)
        search_layout.addStretch()
        
        self.layout.addLayout(search_layout)

    def show_analisis_search(self):
        """Muestra la ventana de búsqueda de análisis unitarios."""
        if not self.analisis_controller:
            self.analisis_controller = AnalisisUnitariosController()
            # Conectar la señal de selección de análisis
            self.analisis_controller.view.analysis_selected.connect(self.on_analisis_selected_from_search)
        
        # Obtener valores de los campos de búsqueda
        codigo = self.codigo_search.text().strip()
        descripcion = self.descripcion_search.text().strip()
        
        # Desconectar temporalmente los eventos de cambio de texto para evitar interferencias
        try:
            self.analisis_controller.view.search_code_input.textChanged.disconnect()
            self.analisis_controller.view.search_desc_input.textChanged.disconnect()
        except:
            pass  # Si no estaban conectadas, ignorar el error
        
        # Limpiar y establecer los valores en campos separados
        self.analisis_controller.view.search_code_input.clear()
        self.analisis_controller.view.search_desc_input.clear()
        
        if codigo:
            self.analisis_controller.view.search_code_input.setText(codigo)
        if descripcion:
            self.analisis_controller.view.search_desc_input.setText(descripcion)
        
        # Reconectar los eventos
        self.analisis_controller.view.search_code_input.textChanged.connect(
            self.analisis_controller.view.apply_filters)
        self.analisis_controller.view.search_desc_input.textChanged.connect(
            self.analisis_controller.view.apply_filters)
        
        # Aplicar filtros manualmente
        self.analisis_controller.view.apply_filters()
        
        # Mostrar la ventana
        self.analisis_controller.view.show()

    def on_analisis_selected_from_search(self, codigo):
        """Oculta la ventana de búsqueda tras seleccionar (la inserción la maneja MainWindow)."""
        if self.analisis_controller:
            self.analisis_controller.view.hide()

    def open_insert_item_dialog(self):
        """Abre un diálogo reutilizando la vista de análisis ya cargada (sin recargar datos)."""
        # Si ya tenemos el controlador inyectado, reusamos su vista moviéndola temporalmente al diálogo
        if getattr(self, 'analisis_controller', None):
            view = self.analisis_controller.view
            original_parent = view.parent()
            original_layout = original_parent.layout() if original_parent else None

            # Quitar del layout original para reparentar
            try:
                if original_layout is not None:
                    original_layout.removeWidget(view)
            except Exception:
                pass

            # ---- Mejorar formato de filas finales: Valor Total, Plazo, Firma ----
            # Solo si existe una hoja de presupuesto cargada (ws_ppto); si no, saltar seguro
            if 'ws_ppto' not in locals():
                # No hay hoja cargada en este contexto; omitir formateos especiales
                ws_ppto = None
            if ws_ppto:
                def _find_row_by_labels(sheet, labels: list[str]):
                    Ls = [_norm(x) for x in labels]
                    for rr in range(1, sheet.max_row + 1):
                        for cc in range(1, min(12, sheet.max_column or 12) + 1):
                            v = sheet.cell(row=rr, column=cc).value
                            if isinstance(v, str) and any(lbl in _norm(v) for lbl in Ls):
                                return rr
                    return None

            header_fill = XLFill(fill_type='solid', start_color='FFEFEFEF', end_color='FFEFEFEF')
            bold_font = XLFont(bold=True)
            medium = XLSide(style='medium', color='FF000000')
            box_border = XLBorder(left=medium, right=medium, top=medium, bottom=medium)

            if ws_ppto:
                # Valor total presupuesto (cabecera y caja de letras)
                vt_row = _find_row_by_labels(ws_ppto, ['VALOR  TOTAL PRESUPUESTO', 'VALOR TOTAL PRESUPUESTO'])
                if vt_row:
                    try:
                        ws_ppto.row_dimensions[vt_row].height = 20
                    except Exception:
                        pass
                    for c in range(2, 7):
                        try:
                            # Asegurar que estilizamos la celda ancla
                            label_cell = _set(ws_ppto, vt_row, c, ws_ppto.cell(row=vt_row, column=c).value)
                            label_cell.font = bold_font
                            label_cell.fill = header_fill
                            label_cell.alignment = XLAlign(horizontal='center', vertical='center')
                            label_cell.border = box_border
                        except Exception:
                            pass
                    # Fila inferior (letras) con borde caja y wrap
                    letras_row = vt_row + 1
                    try:
                        ws_ppto.row_dimensions[letras_row].height = 30
                    except Exception:
                        pass
                    for c in range(2, 7):
                        try:
                            cell = _set(ws_ppto, letras_row, c, ws_ppto.cell(row=letras_row, column=c).value)
                            cell.alignment = XLAlign(horizontal='center', vertical='center', wrap_text=True)
                            cell.border = box_border
                        except Exception:
                            pass

                # Plazo de entrega (caja grande y número a la derecha)
                plazo_row = _find_row_by_labels(ws_ppto, ['PLAZO DE ENTREGA', 'PLAZO DE ENTREGA: (DIAS'])
                if plazo_row:
                    try:
                        ws_ppto.row_dimensions[plazo_row].height = 22
                    except Exception:
                        pass
                    for c in range(3, 6):
                        try:
                            cell = _set(ws_ppto, plazo_row, c, ws_ppto.cell(row=plazo_row, column=c).value)
                            cell.font = bold_font
                            cell.alignment = XLAlign(horizontal='center', vertical='center', wrap_text=True)
                            cell.border = box_border
                        except Exception:
                            pass
                    try:
                        # No reasignar el valor (para evitar sobrescribir el total si el rango está combinado);
                        # solo aplicar formato al ancla existente
                        num_cell = ws_ppto.cell(row=plazo_row, column=6)
                        if isinstance(num_cell, MergedCell):
                            num_cell = _anchor(ws_ppto, plazo_row, 6)
                        num_cell.alignment = XLAlign(horizontal='center', vertical='center')
                        num_cell.border = box_border
                        num_cell.number_format = '0'
                    except Exception:
                        pass

                # Firma del representante legal: línea superior gruesa
                firma_row = _find_row_by_labels(ws_ppto, ['FIRMA DEL REPRESENTANTE LEGAL'])
                if firma_row:
                    try:
                        ws_ppto.row_dimensions[firma_row].height = 18
                    except Exception:
                        pass
                    for c in range(2, 7):
                        try:
                            cell = _set(ws_ppto, firma_row, c, ws_ppto.cell(row=firma_row, column=c).value)
                            cell.alignment = XLAlign(horizontal='center', vertical='center')
                            cell.border = XLBorder(top=medium)
                        except Exception:
                            pass

            dialog = QDialog(self)
            dialog.setWindowTitle("Insertar Ítem - Análisis Unitarios")
            dialog.resize(900, 600)
            dlayout = QVBoxLayout(dialog)
            dlayout.addWidget(view)

            # No conectar señales adicionales: la app principal ya maneja la inserción

            dialog.exec()

            # Restaurar vista al contenedor original
            try:
                dlayout.removeWidget(view)
            except Exception:
                pass
            view.setParent(original_parent)
            if original_layout is not None:
                original_layout.addWidget(view)
            # Nota: no desconectamos _on_select para permitir futuras inserciones rápidas
            return

        # Fallback: si no hay analisis_controller, crear uno (primer uso) y mostrarlo en diálogo
        from controllers.analisis_unitarios_controller import AnalisisUnitariosController
        self.analisis_controller = AnalisisUnitariosController()
        dialog = QDialog(self)
        dialog.setWindowTitle("Insertar Ítem - Análisis Unitarios")
        dialog.resize(900, 600)
        dlayout = QVBoxLayout(dialog)
        dlayout.addWidget(self.analisis_controller.view)
        self.analisis_controller.view.analysis_selected.connect(self.on_analisis_selected_from_search)
        dialog.exec()

    def edit_selected_item_description(self):
        """Permite editar la descripción de la fila seleccionada del presupuesto."""
        selected_items = self.table.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Sin selección", "Seleccione una fila de análisis para modificar su descripción.")
            return
        row = selected_items[0].row()
        code_item = self.table.item(row, 0)
        if not code_item:
            QMessageBox.warning(self, "Error", "La fila seleccionada no es válida.")
            return
        user_role = code_item.data(Qt.ItemDataRole.UserRole)
        if user_role in (None, "", "chapter", "subtotal"):
            QMessageBox.warning(self, "Operación no válida", "Solo se pueden modificar descripciones de filas de análisis (no capítulos ni subtotales).")
            return

        desc_item = self.table.item(row, 1)
        current_desc = desc_item.text() if desc_item else ""
        new_desc, ok = QInputDialog.getText(self, "Modificar Descripción", "Nueva descripción:", text=current_desc)
        if ok and new_desc.strip():
            if not desc_item:
                desc_item = QTableWidgetItem()
                self.table.setItem(row, 1, desc_item)
            desc_item.setText(new_desc.strip())
            desc_item.setToolTip(new_desc.strip())
            self.update_total_presupuesto()

    def create_buttons(self):
        """Crea un conjunto compacto de menús por categoría para reducir saturación visual."""
        button_layout = QHBoxLayout()
        button_layout.setSpacing(8)
        button_layout.setContentsMargins(0, 0, 0, 0)
        
        # Helper para crear botones-menú
        def make_menu_button(title: str, items: list[tuple[str, callable]]):
            btn = QToolButton(self)
            btn.setText(title)
            btn.setPopupMode(QToolButton.ToolButtonPopupMode.InstantPopup)
            btn.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
            menu = QMenu(btn)
            for text, slot in items:
                act = QAction(text, self)
                act.triggered.connect(slot)
                menu.addAction(act)
            # Asegurar que el menú cubra exactamente el ancho del botón al abrirse
            def _resize_menu_to_button():
                try:
                    menu.setFixedWidth(btn.width())
                except Exception:
                    pass
            menu.aboutToShow.connect(_resize_menu_to_button)
            btn.setMenu(menu)
            return btn
        
        # Capítulos
        btn_capitulos = make_menu_button("Capítulos", [
            ("Agregar Capítulo", self.prompt_add_chapter),
            ("Editar Capítulo", self.prompt_edit_chapter),
            ("Eliminar Capítulo", self.prompt_delete_chapter),
        ])
        
        # Ítems
        btn_items = make_menu_button("Ítems", [
            ("Insertar Ítem", self.open_insert_item_dialog),
            ("Modificar Descripción", self.edit_selected_item_description),
            ("Eliminar Ítem", self.delete_selected_row),
        ])
        
        # Análisis
        btn_analisis = make_menu_button("Análisis", [
            ("Buscar Análisis (fila)", self.open_match_dialog_for_selected),
            ("Editar Análisis (fila)", self.edit_selected_analysis),
        ])
        
        # Importar / Exportar
        btn_io = make_menu_button("Importar/Exportar", [
            ("Importar CSV", self.import_csv),
            ("Importar por Texto", self.open_import_text_dialog),
            ("Importar IFC (Materiales)", self.open_ifc_materials_dialog),
            ("Exportar CSV", self.export_csv),
            ("Exportar Excel", self.export_excel_from_scratch),
        ])
        
        # Añadir en orden lógico
        button_layout.addWidget(btn_capitulos, 1)
        button_layout.addWidget(btn_items, 1)
        button_layout.addWidget(btn_analisis, 1)
        button_layout.addWidget(btn_io, 1)
        
        self.layout.addLayout(button_layout)

    def open_ifc_materials_dialog(self):
        try:
            from .ifc_materials_dialog import IFCMaterialsDialog
        except Exception as e:
            QMessageBox.critical(self, "Error", f"No se pudo cargar el diálogo IFC:\n{e}")
            return
        dlg = IFCMaterialsDialog(self)
        dlg.exec()

    def open_import_text_dialog(self, prefill_rows=None, append: bool = False):
        dialog = ImportarPorTextoDialog(self, prefill_rows=prefill_rows)
        if dialog.exec():
            rows = dialog.result_rows()
            if not rows:
                return
            # Insertar filas como capítulos o análisis según ITEM
            self.table.blockSignals(True)
            try:
                # Limpiar presupuesto actual y contadores solo si no se desea anexar
                if not append:
                    self.table.setRowCount(0)
                    self.chapter_counter = 0
                for entry in rows:
                    item_str = (entry.get('item') or '').strip()
                    desc = (entry.get('descripcion') or '').strip()
                    und = (entry.get('unidad') or '').strip()
                    cantidad = entry.get('cantidad') or 1.0

                    if not desc:
                        continue

                    # Capítulo: número entero (1, 2, 3)
                    is_chapter = False
                    if item_str:
                        try:
                            # válido si es entero con solo dígitos
                            is_chapter = item_str.isdigit()
                        except Exception:
                            is_chapter = False

                    if is_chapter:
                        self.add_chapter_row(desc, trigger_rebuild=False)
                        continue

                    # Análisis: subnúmero como 1.1, 2.03, 3.1.1 (se acepta como texto)
                    # Insertamos fila de análisis vacía con unidad/cantidad; costo unitario en 0 y pendiente de match
                    row_idx = self.table.rowCount()
                    self.table.insertRow(row_idx)
                    # Columna 0: item (se renumerará después), guardamos marcador para diferenciar de capítulos
                    code_item = QTableWidgetItem(item_str or "...")
                    self.table.setItem(row_idx, 0, code_item)

                    # Columna 1: descripción
                    desc_item = QTableWidgetItem(desc)
                    desc_item.setToolTip(desc)
                    self.table.setItem(row_idx, 1, desc_item)

                    # Columna 2: unidad
                    und_item = QTableWidgetItem(und.upper())
                    und_item.setFlags(und_item.flags() | Qt.ItemFlag.ItemIsEditable)
                    self.table.setItem(row_idx, 2, und_item)

                    # Columna 3: cantidad
                    try:
                        qty_val = float(cantidad)
                    except Exception:
                        qty_val = 1.0
                    qty_item = QTableWidgetItem(str(qty_val))
                    qty_item.setFlags(qty_item.flags() | Qt.ItemFlag.ItemIsEditable)
                    self.table.setItem(row_idx, 3, qty_item)

                    # Columna 4: costo unitario (se completará al emparejar con análisis de BD)
                    cu_item = QTableWidgetItem("$0.00")
                    self.table.setItem(row_idx, 4, cu_item)

                    # Columna 5: total
                    ct_item = QTableWidgetItem("$0.00")
                    self.table.setItem(row_idx, 5, ct_item)

                # Renumerar y recalcular subtotales, sin tocar costos unitarios
                self.rebuild_table_safe()
            finally:
                self.table.blockSignals(False)

            # Opcional: lanzar buscador para completar costos unitarios ahora
            try:
                ask = QMessageBox.question(
                    self,
                    "Buscar Análisis",
                    "¿Desea buscar coincidencias en la base de datos para los ítems importados y seleccionar costos unitarios ahora?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                    QMessageBox.StandardButton.Yes,
                )
            except Exception:
                ask = QMessageBox.StandardButton.No

            if ask == QMessageBox.StandardButton.Yes:
                # Recorrer ítems de análisis (no capítulos ni subtotales)
                interrupted = False
                for r in range(self.table.rowCount()):
                    if interrupted:
                        break
                    code_item = self.table.item(r, 0)
                    if not code_item:
                        continue
                    role = code_item.data(Qt.ItemDataRole.UserRole)
                    if role in ["chapter", "subtotal"]:
                        continue

                    desc = self.table.item(r, 1).text() if self.table.item(r, 1) else ""
                    und = self.table.item(r, 2).text() if self.table.item(r, 2) else ""
                    if not desc:
                        continue

                    dlg = AnalisisMatchDialog(desc, und, parent=self)
                    result = dlg.exec()
                    # Si el usuario pulsó "Detener proceso" en el diálogo, abortamos el resto
                    try:
                        if not result and getattr(dlg, 'was_aborted', lambda: False)():
                            interrupted = True
                            break
                    except Exception:
                        pass
                    if result:
                        sel = dlg.selected_analysis()
                        if sel:
                            try:
                                self.table.blockSignals(True)
                                # Guardar código en UserRole de la columna 0
                                if not code_item:
                                    code_item = QTableWidgetItem("...")
                                    self.table.setItem(r, 0, code_item)
                                code_item.setData(Qt.ItemDataRole.UserRole, sel['codigo'])

                                # Unidad
                                und_item = self.table.item(r, 2)
                                if not und_item:
                                    und_item = QTableWidgetItem("")
                                    self.table.setItem(r, 2, und_item)
                                und_item.setText((sel['unidad'] or '').upper())

                                # Costo unitario
                                cu_item = self.table.item(r, 4)
                                if not cu_item:
                                    cu_item = QTableWidgetItem()
                                    self.table.setItem(r, 4, cu_item)
                                cu_item.setText(f"${sel['costo_unitario']:,.2f}")

                                # Recalcular totales
                                self.update_row_total(r)
                            finally:
                                self.table.blockSignals(False)

                # Actualizar totales generales y subtotales después del proceso
                self.update_total_presupuesto()

    def open_match_dialog_for_selected(self):
        """Abre el buscador de análisis por unidad y descripción para la fila seleccionada."""
        selected_items = self.table.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Sin selección", "Seleccione una fila de análisis para buscar coincidencias.")
            return
        row = selected_items[0].row()

        # No permitir capítulos ni subtotales
        code_item = self.table.item(row, 0)
        if not code_item or (code_item.data(Qt.ItemDataRole.UserRole) in ["chapter", "subtotal"]):
            QMessageBox.warning(self, "Operación no válida", "Seleccione una fila de análisis (no capítulo ni subtotal).")
            return

        desc = self.table.item(row, 1).text() if self.table.item(row, 1) else ""
        und = self.table.item(row, 2).text() if self.table.item(row, 2) else ""

        if not desc:
            QMessageBox.warning(self, "Datos incompletos", "La fila seleccionada no tiene descripción.")
            return

        dialog = AnalisisMatchDialog(desc, und, parent=self)
        if dialog.exec():
            sel = dialog.selected_analysis()
            if not sel:
                return
            # Aplicar: código, unidad y costo unitario
            # 1) Guardar código original en UserRole para la fila
            try:
                self.table.blockSignals(True)
                code_item = self.table.item(row, 0)
                if not code_item:
                    code_item = QTableWidgetItem("...")
                    self.table.setItem(row, 0, code_item)
                code_item.setData(Qt.ItemDataRole.UserRole, sel['codigo'])
                # 2) Actualizar unidad si está vacía o distinta
                und_item = self.table.item(row, 2)
                if not und_item:
                    und_item = QTableWidgetItem("")
                    self.table.setItem(row, 2, und_item)
                und_item.setText((sel['unidad'] or '').upper())
                # 3) Establecer costo unitario
                cu_item = self.table.item(row, 4)
                if not cu_item:
                    cu_item = QTableWidgetItem()
                    self.table.setItem(row, 4, cu_item)
                cu_item.setText(f"${sel['costo_unitario']:,.2f}")
                # 4) Recalcular total de la fila y totales
                self.update_row_total(row)
                self.update_total_presupuesto()
            finally:
                self.table.blockSignals(False)

    def prompt_add_chapter(self):
        """Pide al usuario el nombre del capítulo y lo agrega."""
        text, ok = QInputDialog.getText(self, 'Agregar Capítulo', 'Nombre del capítulo:')
        if ok and text:
            self.add_chapter_row(text)

    def prompt_edit_chapter(self):
        """Permite editar el nombre del capítulo seleccionado."""
        selected_items = self.table.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Advertencia", "Por favor, seleccione un capítulo para editar.")
            return
        
        # Obtener la fila seleccionada
        row = selected_items[0].row()
        item = self.table.item(row, 0)
        
        # Verificar que sea un capítulo
        if not item or item.data(Qt.ItemDataRole.UserRole) != 'chapter':
            QMessageBox.warning(self, "Advertencia", "Por favor, seleccione una fila de capítulo para editar.")
            return
        
        # Extraer el nombre actual del capítulo (sin el número)
        current_text = item.text()
        try:
            current_name = current_text.split('.', 1)[1].strip()
        except IndexError:
            current_name = current_text
        
        # Pedir el nuevo nombre
        new_name, ok = QInputDialog.getText(
            self, 
            'Editar Capítulo', 
            'Nuevo nombre del capítulo:',
            text=current_name
        )
        
        if ok and new_name.strip():
            # Actualizar el nombre del capítulo
            chapter_number = item.data(Qt.ItemDataRole.UserRole + 1)
            if chapter_number:
                item.setText(f"{chapter_number}. {new_name.strip().upper()}")
            else:
                item.setText(new_name.strip().upper())
            
            QMessageBox.information(self, "Capítulo actualizado", "El nombre del capítulo ha sido actualizado correctamente.")

    def prompt_delete_chapter(self):
        """Elimina el capítulo seleccionado junto con todos sus análisis asociados."""
        selected_items = self.table.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Advertencia", "Por favor, seleccione un capítulo para eliminar.")
            return
        
        # Obtener la fila seleccionada
        row = selected_items[0].row()
        item = self.table.item(row, 0)
        
        # Verificar que sea un capítulo
        if not item or item.data(Qt.ItemDataRole.UserRole) != 'chapter':
            QMessageBox.warning(self, "Advertencia", "Por favor, seleccione una fila de capítulo para eliminar.")
            return
        
        # Extraer el nombre del capítulo
        chapter_name = item.text()
        
        # Contar cuántos análisis tiene el capítulo
        analysis_count = 0
        chapter_end_row = self.table.rowCount()
        
        # Buscar el final del capítulo
        for i in range(row + 1, self.table.rowCount()):
            next_item = self.table.item(i, 0)
            if next_item:
                if next_item.data(Qt.ItemDataRole.UserRole) == 'chapter':
                    chapter_end_row = i
                    break
                elif next_item.data(Qt.ItemDataRole.UserRole) not in ['subtotal']:
                    analysis_count += 1
        
        # Pedir confirmación
        if analysis_count > 0:
            confirmacion = QMessageBox.question(
                self,
                "Confirmar eliminación de capítulo",
                f"¿Está seguro de eliminar el capítulo '{chapter_name}' y todos sus {analysis_count} análisis asociados?\n\n"
                f"Esta acción no se puede deshacer.",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
        else:
            confirmacion = QMessageBox.question(
                self,
                "Confirmar eliminación de capítulo",
                f"¿Está seguro de eliminar el capítulo '{chapter_name}'?\n\n"
                f"Este capítulo no tiene análisis asociados.",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
        
        if confirmacion == QMessageBox.StandardButton.Yes:
            # Eliminar todas las filas del capítulo (desde el final hacia el principio)
            rows_to_delete = []
            
            # Recopilar todas las filas que pertenecen al capítulo
            for i in range(row, chapter_end_row):
                if i < self.table.rowCount():
                    rows_to_delete.append(i)
            
            # Eliminar las filas desde el final hacia el principio para no afectar los índices
            for delete_row in reversed(rows_to_delete):
                if delete_row < self.table.rowCount():
                    self.table.removeRow(delete_row)
            
            # Renumerar y actualizar totales
            self.rebuild_table()
            
            if analysis_count > 0:
                QMessageBox.information(
                    self, 
                    "Capítulo eliminado", 
                    f"El capítulo '{chapter_name}' y sus {analysis_count} análisis han sido eliminados correctamente."
                )
            else:
                QMessageBox.information(
                    self, 
                    "Capítulo eliminado", 
                    f"El capítulo '{chapter_name}' ha sido eliminado correctamente."
                )

    def add_chapter_row(self, chapter_name, trigger_rebuild=True):
        """Agrega una fila de capítulo al final de la tabla.
        Si trigger_rebuild es False, no ejecuta la reconstrucción completa (útil durante importaciones masivas)."""
        row = self.table.rowCount()
        self.table.insertRow(row)
        
        item = QTableWidgetItem(chapter_name.upper())  # El número se asignará al renumerar
        item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled)  # No editable
        
        # Estilo para el capítulo
        font = QFont()
        font.setBold(True)
        item.setFont(font)
        item.setBackground(QColor("#e0e0e0"))
        item.setData(Qt.ItemDataRole.UserRole, 'chapter')
        item.setData(Qt.ItemDataRole.UserRole + 1, 0)  # El número se asignará al renumerar

        self.table.setItem(row, 0, item)
        self.table.setSpan(row, 0, 1, self.table.columnCount())
        
        # Solo reconstruir inmediatamente si no estamos importando en bloque
        if trigger_rebuild:
            self.rebuild_table()

    def add_ifc_rebar_analysis(self, kg_total: float):
        """Agrega directamente el análisis de ACERO DE REFUERZO (08-08-11)
        con la cantidad total en kilogramos extraída del IFC, sin pasar por búsqueda."""
        try:
            if not kg_total or kg_total <= 0:
                return
            self.table.blockSignals(True)

            # Asegurar que exista el capítulo ACERO (IFC)
            chapter_exists = False
            for r in range(self.table.rowCount()):
                it = self.table.item(r, 0)
                if it and it.data(Qt.ItemDataRole.UserRole) == 'chapter':
                    if (it.text() or '').strip().upper().startswith('ACERO (IFC)'):
                        chapter_exists = True
                        break
            if not chapter_exists:
                self.add_chapter_row('ACERO (IFC)', trigger_rebuild=False)

            # Insertar fila de análisis al final
            row = self.table.rowCount()
            self.table.insertRow(row)

            # Columna 0: item (se renumerará). Guardamos el código en UserRole
            code_item = QTableWidgetItem('...')
            code_item.setData(Qt.ItemDataRole.UserRole, '08-08-11')
            self.table.setItem(row, 0, code_item)

            # Columna 1: descripción
            desc_text = 'ACERO DE REFUERZO (08-08-11)'
            desc_item = QTableWidgetItem(desc_text)
            desc_item.setToolTip(desc_text)
            self.table.setItem(row, 1, desc_item)

            # Unidad y costo unitario desde BD (si existe)
            und_text = 'KG'
            cu_value = 0.0
            try:
                session = SessionLocal()
                try:
                    a = session.query(AnalisisUnitario).filter(AnalisisUnitario.codigo == '08-08-11').first()
                    if a:
                        try:
                            und_text = (a.unidad or und_text).upper()
                        except Exception:
                            pass
                        try:
                            cu_value = float(getattr(a, 'total_calculado', None) or getattr(a, 'total', 0.0) or 0.0)
                        except Exception:
                            cu_value = 0.0
                finally:
                    session.close()
            except Exception:
                pass

            # Columna 2: unidad
            und_item = QTableWidgetItem(und_text)
            und_item.setFlags(und_item.flags() | Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(row, 2, und_item)

            # Columna 3: cantidad
            qty_item = QTableWidgetItem(f"{kg_total:.2f}")
            qty_item.setFlags(qty_item.flags() | Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(row, 3, qty_item)

            # Columna 4: costo unitario
            cu_item = QTableWidgetItem(f"${cu_value:,.2f}")
            self.table.setItem(row, 4, cu_item)

            # Calcular total y renumerar
            self.update_row_total(row)
            self.rebuild_table_safe()
        finally:
            self.table.blockSignals(False)

    def delete_selected_row(self):
        """Elimina la fila seleccionada del presupuesto."""
        selected_items = self.table.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Advertencia", "Por favor, seleccione una fila para eliminar.")
            return
        
        # Obtener la fila seleccionada (tomamos el primer ítem seleccionado)
        row = selected_items[0].row()
        
        # Verificar el tipo de fila seleccionada
        codigo_item = self.table.item(row, 0)
        if codigo_item:
            user_role = codigo_item.data(Qt.ItemDataRole.UserRole)
            
            # Bloquear eliminación de capítulos
            if user_role == 'chapter':
                QMessageBox.warning(
                    self, 
                    "Operación no permitida", 
                    "No se puede eliminar un capítulo usando 'Eliminar Item'.\n\n"
                    "Use el botón 'Eliminar Capítulo' para eliminar capítulos completos."
                )
                return
            
            # Bloquear eliminación de subtotales
            if user_role == 'subtotal':
                QMessageBox.warning(
                    self, 
                    "Operación no permitida", 
                    "No se puede eliminar un subtotal.\n\n"
                    "Los subtotales se calculan automáticamente."
                )
                return
            
            # Solo permitir eliminación de análisis individuales
            codigo = codigo_item.text()
            descripcion_item = self.table.item(row, 1)
            descripcion = descripcion_item.text() if descripcion_item else ""
            
            confirmacion = QMessageBox.question(
                self,
                "Confirmar eliminación",
                f"¿Está seguro de eliminar el análisis '{codigo} - {descripcion}'?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            
            if confirmacion == QMessageBox.StandardButton.Yes:
                self.table.removeRow(row)
                self.rebuild_table()
                QMessageBox.information(self, "Eliminado", "El análisis ha sido eliminado correctamente.")
        else:
            QMessageBox.warning(self, "Error", "No se puede determinar el tipo de fila seleccionada.")

    def create_table(self):
        """Crea la tabla para mostrar los análisis unitarios del presupuesto."""
        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels([
            "Item", "Descripción", "Unidad", 
            "Cantidad", "Costo Unitario", "Costo Total"
        ])
        
        # Configurar el ancho de las columnas
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)  # Item
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)          # Descripción
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)  # Unidad
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)  # Cantidad
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)  # Costo Unitario
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)  # Costo Total
        
        self.layout.addWidget(self.table)
        
        # Agregar fila para el total
        self.total_label = QLabel("Total del Presupuesto: $0.00")
        self.total_label.setAlignment(Qt.AlignmentFlag.AlignRight)
        self.total_label.setStyleSheet("font-size: 16px; font-weight: bold; padding: 10px;")
        self.layout.addWidget(self.total_label)
        
        # Desglose AIU
        self.breakdown_label = QLabel("")
        self.breakdown_label.setAlignment(Qt.AlignmentFlag.AlignRight)
        self.breakdown_label.setStyleSheet("font-size: 13px; padding-right: 10px;")
        self.layout.addWidget(self.breakdown_label)
        
        # Conectar el evento de cambio de celda
        self.table.itemChanged.connect(self.on_cell_changed)
        self.table.cellClicked.connect(self.on_cell_clicked)

    def on_cell_clicked(self, row, column):
        """Maneja los clics en las celdas, detectando Shift+Click para editar."""
        modifiers = QApplication.keyboardModifiers()
        if modifiers == Qt.KeyboardModifier.ShiftModifier:
            item = self.table.item(row, 0)  # El código del análisis está en la data del item
            if item and item.data(Qt.ItemDataRole.UserRole) not in ['chapter', 'subtotal']:
                analisis_code = item.data(Qt.ItemDataRole.UserRole)
                if analisis_code:
                    self.analysis_edit_requested.emit(analisis_code)

    def add_analisis(self, analisis_data):
        """Agrega un análisis unitario a la tabla bajo el capítulo seleccionado o el último."""
        
        # Verificar que los datos del análisis estén completos
        required_fields = ['codigo', 'descripcion', 'unidad', 'costo_unitario']
        for field in required_fields:
            if field not in analisis_data:
                QMessageBox.warning(self, "Error", f"Datos incompletos del análisis: falta {field}")
                return
        
        # Determinar el capítulo y la posición de inserción
        target_chapter_row = -1
        
        current_row = self.table.currentRow()
        if current_row == -1: # Nada seleccionado, usar el último capítulo
            start_row = self.table.rowCount() - 1
        else:
            start_row = current_row

        # Buscar hacia atrás el capítulo al que pertenece la selección
        for i in range(start_row, -1, -1):
            item = self.table.item(i, 0)
            if item and item.data(Qt.ItemDataRole.UserRole) == 'chapter':
                target_chapter_row = i
                break
                
        if target_chapter_row == -1:
            QMessageBox.warning(self, "Advertencia", "Por favor, agregue un capítulo antes de añadir un análisis.")
            return

        # Encontrar el final del capítulo para insertar la nueva fila
        insertion_row = self.table.rowCount() # Por defecto, al final
        for i in range(target_chapter_row + 1, self.table.rowCount()):
            item = self.table.item(i, 0)
            if item:
                # Si encontramos otro capítulo, insertar antes de él
                if item.data(Qt.ItemDataRole.UserRole) == 'chapter':
                    insertion_row = i
                    break
                # Si encontramos un subtotal, insertar antes de él (final del capítulo actual)
                elif item.data(Qt.ItemDataRole.UserRole) == 'subtotal':
                    insertion_row = i
                    break

        # Insertar la nueva fila
        self.table.insertRow(insertion_row)
        
        # Crear y configurar todos los QTableWidgetItem primero
        items = []
        for col in range(6):
            item = QTableWidgetItem()
            if col not in (2, 3):  # Columnas 2 (Unidad) y 3 (Cantidad) editables
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            items.append(item)
            self.table.setItem(insertion_row, col, item)
        
        # Establecer los valores
        items[0].setText("...") # Temporal, se renumerará
        items[0].setData(Qt.ItemDataRole.UserRole, analisis_data['codigo']) # Guardar código original
        
        descripcion = analisis_data['descripcion']
        items[1].setText(descripcion)
        items[1].setToolTip(descripcion)
        
        items[2].setText(analisis_data['unidad'])
        items[3].setText('1')  # Cantidad por defecto
        
        # Establecer el costo unitario asegurándose de que sea un número válido
        try:
            costo_unitario = float(analisis_data['costo_unitario'])
            items[4].setText(f"${costo_unitario:,.2f}")
        except (ValueError, TypeError):
            items[4].setText("$0.00")
        
        # El total se calculará automáticamente
        items[5].setText('$0.00')
        
        # Renumerar todo y actualizar totales
        self.renumber_items()
        self.update_row_total(insertion_row)
        self.update_total_presupuesto()

    def on_cell_changed(self, item):
        """Maneja los cambios en las celdas de la tabla."""
        # Solo procesar cambios en la columna Cantidad (índice 3)
        if not item or item.column() != 3:
            return
            
        # Guardar la fila antes de cualquier procesamiento que pueda invalidar el item
        row = item.row()
        
        # Verificar que el item sigue siendo válido
        try:
            text = item.text().strip()
        except RuntimeError:
            # El item ha sido eliminado, salir sin procesar
            return
            
        try:
            if not text:  # Si está vacío, establecer en 1
                # Verificar que el item sigue siendo válido antes de modificarlo
                try:
                    item.setText('1')
                except RuntimeError:
                    return
                cantidad = 1.0
            else:
                cantidad = float(text)
                if cantidad < 0:
                    raise ValueError("La cantidad no puede ser negativa")
            
            # Bloquear señales para evitar recursión
            self.table.blockSignals(True)
            
            # Verificar que la fila sigue siendo válida después de bloquear señales
            if row < self.table.rowCount():
                self.update_row_total(row)
            self.update_total_presupuesto()
            
            self.table.blockSignals(False)
            
        except ValueError:
            self.table.blockSignals(True)
            
            # Verificar que el item sigue siendo válido antes de modificarlo
            try:
                item.setText('1')
                if row < self.table.rowCount():
                    self.update_row_total(row)
                self.update_total_presupuesto()
            except RuntimeError:
                pass  # El item ha sido eliminado, no hacer nada
            
            self.table.blockSignals(False)
            QMessageBox.warning(self, "Error", "Por favor ingrese un número válido positivo")

    def update_row_total(self, row):
        """Actualiza el costo total de una fila basado en la cantidad y el costo unitario."""
        # Verificar que la fila existe
        if row < 0 or row >= self.table.rowCount():
            return
            
        try:
            cantidad_item = self.table.item(row, 3)
            costo_unitario_item = self.table.item(row, 4)
            
            if not cantidad_item or not costo_unitario_item:
                return
            
            # Obtener los valores
            cantidad_text = cantidad_item.text().replace(',', '').strip()
            costo_text = costo_unitario_item.text().replace('$', '').replace(',', '').strip()
            
            # Convertir a números, usar valores por defecto si hay error
            cantidad = float(cantidad_text) if cantidad_text else 1.0
            costo_unitario = float(costo_text) if costo_text else 0.0
            
            total = cantidad * costo_unitario
            
            # Actualizar el total
            total_item = self.table.item(row, 5)
            if not total_item:
                total_item = QTableWidgetItem()
                self.table.setItem(row, 5, total_item)
            
            total_item.setText(f"${total:,.2f}")
                
        except (ValueError, TypeError, RuntimeError):
            # Si hay cualquier error, establecer como Error
            try:
                total_item = self.table.item(row, 5)
                if not total_item:
                    total_item = QTableWidgetItem()
                    self.table.setItem(row, 5, total_item)
                total_item.setText("Error")
            except (RuntimeError, IndexError):
                pass  # Si no se puede ni siquiera crear el item, ignorar

    def update_total_presupuesto(self):
        """Recalcula y actualiza el costo total del presupuesto y los subtotales por capítulo."""
        # Bloquear señales para evitar llamadas recursivas
        self.table.blockSignals(True)
        
        try:
            # Primero, eliminar todas las filas de subtotal existentes
            self.remove_subtotal_rows()
            
            # Recopilar información sobre capítulos y calcular subtotales
            total_presupuesto = 0.0
            current_chapter_start = -1
            current_chapter_total = 0.0
            rows_to_insert_subtotals = []
            
            row = 0
            while row < self.table.rowCount():
                item = self.table.item(row, 0)
                
                if item and item.data(Qt.ItemDataRole.UserRole) == 'chapter':
                    # Si había un capítulo anterior, guardar su información para insertar subtotal
                    if current_chapter_start != -1 and current_chapter_total > 0:
                        rows_to_insert_subtotals.append({
                            'position': row,
                            'total': current_chapter_total
                        })
                    
                    # Iniciar nuevo capítulo
                    current_chapter_start = row
                    current_chapter_total = 0.0
                    
                elif item and item.data(Qt.ItemDataRole.UserRole) not in ['chapter', 'subtotal']:
                    # Es una fila de análisis, calcular su total y sumarlo
                    try:
                        cantidad_item = self.table.item(row, 3)
                        costo_unitario_item = self.table.item(row, 4)
                        
                        if cantidad_item and costo_unitario_item:
                            try:
                                cantidad_text = cantidad_item.text().replace(',', '').strip()
                                costo_text = costo_unitario_item.text().replace('$', '').replace(',', '').strip()
                                
                                # Solo actualizar el total si ya hay un costo unitario
                                if costo_text:  # Solo si hay un costo unitario establecido
                                    cantidad = float(cantidad_text) if cantidad_text else 1.0
                                    costo_unitario = float(costo_text) if costo_text else 0.0
                                    
                                    total_fila = cantidad * costo_unitario
                                    
                                    # Actualizar el total de la fila
                                    total_item = self.table.item(row, 5)
                                    if not total_item:
                                        total_item = QTableWidgetItem()
                                        self.table.setItem(row, 5, total_item)
                                    
                                    total_item.setText(f"${total_fila:,.2f}")
                                    
                                    # Sumarlo al capítulo actual y al total general
                                    current_chapter_total += total_fila
                                    total_presupuesto += total_fila
                                
                            except (ValueError, RuntimeError):
                                # Error en conversión, marcar como error pero continuar
                                total_item = self.table.item(row, 5)
                                if not total_item:
                                    total_item = QTableWidgetItem()
                                    self.table.setItem(row, 5, total_item)
                                total_item.setText("Error")
                    except (RuntimeError, AttributeError):
                        pass  # Ignorar filas problemáticas
                
                row += 1
            
            # Agregar subtotal para el último capítulo si existe
            if current_chapter_start != -1 and current_chapter_total > 0:
                rows_to_insert_subtotals.append({
                    'position': self.table.rowCount(),
                    'total': current_chapter_total
                })
            
            # Insertar subtotales desde el final hacia el principio para no afectar los índices
            for subtotal_info in reversed(rows_to_insert_subtotals):
                self.insert_subtotal_row(subtotal_info['position'], subtotal_info['total'])
            
            # Guardar costo directo
            self.direct_cost_total = total_presupuesto

            # Si existe AIU calculado con profesionales, incluirlo
            admin_total = getattr(self, 'admin_cost_total', 0.0)
            grand_total = total_presupuesto + admin_total

            if admin_total > 0:
                # Mostrar resumen compacto: Costo directo, Indirecto (AIU) y Total
                text = (
                    f"Costo Directo: ${total_presupuesto:,.2f}   "
                    f"AIU: ${admin_total:,.2f}   "
                    f"Total Presupuesto: ${grand_total:,.2f}"
                )
            else:
                text = f"Total del Presupuesto: ${total_presupuesto:,.2f}"

            self.total_label.setText(text)
            
            # Update breakdown label
            if admin_total > 0 and hasattr(self, 'aiu_breakdown'):
                bd = self.aiu_breakdown
                lines = [
                    f"Administración: ${bd['admin']:,.2f}",
                    f"Imprevistos ({bd['imprev_pct']:.2f}%): ${bd['imprev']:,.2f}",
                    f"Utilidad ({bd['util_pct']:.2f}%): ${bd['util']:,.2f}",
                    f"IVA Utilidad ({bd['iva_pct']:.2f}%): ${bd['iva']:,.2f}",
                    f"Total Costos Indirectos: ${bd['total_aiu']:,.2f}"
                ]
                self.breakdown_label.setText("<br>".join(lines))
            else:
                self.breakdown_label.setText("")
            
        finally:
            # Siempre desbloquear señales
            self.table.blockSignals(False)

    def remove_subtotal_rows(self):
        """Elimina todas las filas de subtotal existentes."""
        rows_to_remove = []
        for row in range(self.table.rowCount()):
            try:
                item = self.table.item(row, 0)
                if item and item.data(Qt.ItemDataRole.UserRole) == 'subtotal':
                    rows_to_remove.append(row)
            except RuntimeError:
                # Item eliminado, omitir
                continue
        
        # Eliminar de abajo hacia arriba para no afectar los índices
        for row in reversed(rows_to_remove):
            try:
                if row < self.table.rowCount():
                    self.table.removeRow(row)
            except (RuntimeError, IndexError):
                # Fila ya eliminada o inválida, continuar
                continue

    def insert_subtotal_row(self, position, subtotal_amount):
        """Inserta una fila de subtotal en la posición especificada."""
        self.table.insertRow(position)
        
        # Crear item de subtotal
        subtotal_item = QTableWidgetItem(f"SUBTOTAL")
        subtotal_item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled)  # No editable
        
        # Estilo para el subtotal
        font = QFont()
        font.setBold(True)
        subtotal_item.setFont(font)
        subtotal_item.setBackground(QColor("#f0f0f0"))
        subtotal_item.setData(Qt.ItemDataRole.UserRole, 'subtotal')
        
        self.table.setItem(position, 0, subtotal_item)
        
        # Crear item del valor del subtotal en la columna "Costo Total"
        valor_item = QTableWidgetItem(f"${subtotal_amount:,.2f}")
        valor_item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled)  # No editable
        valor_item.setFont(font)
        valor_item.setBackground(QColor("#f0f0f0"))
        self.table.setItem(position, 5, valor_item)
        
        # Crear items vacíos para las otras columnas
        for col in [1, 2, 3, 4]:
            empty_item = QTableWidgetItem("")
            empty_item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled)  # No editable
            empty_item.setBackground(QColor("#f0f0f0"))
            self.table.setItem(position, col, empty_item)

    def export_csv(self):
        """Exporta los datos del presupuesto a un archivo CSV incluyendo datos completos de AIU."""
        filePath, _ = QFileDialog.getSaveFileName(self, "Exportar CSV", "", "CSV Files (*.csv)")
        if not filePath:
            return

        with open(filePath, 'w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            
            # Escribir header - sin la columna "Costo Total"
            writer.writerow(['Item', 'Descripción', 'Unidad', 'Cantidad', 'Costo Unitario', 'Código Análisis', 'Tipo'])

            for row in range(self.table.rowCount()):
                # Obtener datos de la fila
                item_text = self.table.item(row, 0).text() if self.table.item(row, 0) else ""
                desc_text = self.table.item(row, 1).text() if self.table.item(row, 1) else ""
                
                # Verificar si es una fila de subtotal y omitirla
                if ('SUBTOTAL' in item_text.upper() or 
                    'SUBTOTAL' in desc_text.upper() or
                    (not item_text and not desc_text)):
                    continue
                
                # Escribir datos de la fila
                row_data = []
                for col in range(5):  # Solo las primeras 5 columnas (sin Costo Total)
                    item = self.table.item(row, col)
                    if item:
                        row_data.append(item.text())
                    else:
                        row_data.append("")
                
                # Agregar el código del análisis y tipo
                item_with_code = self.table.item(row, 0)
                if item_with_code and item_with_code.data(Qt.ItemDataRole.UserRole):
                    codigo_analisis = item_with_code.data(Qt.ItemDataRole.UserRole)
                    tipo = codigo_analisis if codigo_analisis in ['chapter', 'subtotal'] else 'analysis'
                else:
                    codigo_analisis = ""
                    tipo = ""
                
                row_data.append(codigo_analisis)
                row_data.append(tipo)
                writer.writerow(row_data)

            # ----- Bloque AIU COMPLETO -----
            writer.writerow([])
            writer.writerow(['=== DATOS AIU ==='])
            
            # Obtener datos de AIU si existen
            if hasattr(self, 'aiu_breakdown') and self.aiu_breakdown:
                aiu_data = self.aiu_breakdown
                
                # Datos principales de AIU
                writer.writerow(['AIU_ADMIN', str(aiu_data.get('admin', 0.0))])
                writer.writerow(['AIU_SUB_TOTAL', str(aiu_data.get('sub_total', 0.0))])
                writer.writerow(['AIU_IMPREV', str(aiu_data.get('imprev', 0.0))])
                writer.writerow(['AIU_UTIL', str(aiu_data.get('util', 0.0))])
                writer.writerow(['AIU_IVA', str(aiu_data.get('iva', 0.0))])
                writer.writerow(['AIU_IMPREV_PCT', str(aiu_data.get('imprev_pct', 0.0))])
                writer.writerow(['AIU_UTIL_PCT', str(aiu_data.get('util_pct', 0.0))])
                writer.writerow(['AIU_IVA_PCT', str(aiu_data.get('iva_pct', 0.0))])
                writer.writerow(['AIU_TOTAL', str(aiu_data.get('total_aiu', 0.0))])
                
                # Sub-items detallados (profesionales, oficina, pólizas, estampillas)
                if 'sub_items' in aiu_data:
                    writer.writerow([])
                    writer.writerow(['=== SUB-ITEMS AIU ==='])
                    for i, sub_item in enumerate(aiu_data['sub_items']):
                        writer.writerow([
                            f'SUB_ITEM_{i}',
                            sub_item.get('concepto', ''),
                            str(sub_item.get('pct', 0.0)),
                            str(sub_item.get('valor', 0.0))
                        ])
            
            writer.writerow([])  # línea en blanco
            writer.writerow(["COSTO DIRECTO", f"{self.direct_cost_total}"])

            if hasattr(self, 'aiu_breakdown'):
                bd = self.aiu_breakdown
                writer.writerow(["ADMINISTRACIÓN", f"{bd['admin']}", f"{bd['admin'] / self.direct_cost_total * 100:.2f}%"])
                writer.writerow(["IMPREVISTOS", f"{bd['imprev']}", f"{bd['imprev_pct']:.2f}%"])
                writer.writerow(["UTILIDAD", f"{bd['util']}", f"{bd['util_pct']:.2f}%"])
                writer.writerow(["IVA UTILIDAD", f"{bd['iva']}", f"{bd['iva_pct']:.2f}%"])
                writer.writerow(["TOTAL COSTOS INDIRECTOS", f"{bd['total_aiu']}"])
                total_pres = self.direct_cost_total + bd['total_aiu']
                writer.writerow(["TOTAL PRESUPUESTO", f"{total_pres}"])
            else:
                writer.writerow(["ADMINISTRACIÓN", "$0.00"])
                writer.writerow(["IMPREVISTOS", "$0.00"])
                writer.writerow(["UTILIDAD", "$0.00"])
                writer.writerow(["IVA UTILIDAD", "$0.00"])
                writer.writerow(["TOTAL COSTOS INDIRECTOS", "$0.00"])
                writer.writerow(["TOTAL PRESUPUESTO", f"{self.direct_cost_total}"])

        QMessageBox.information(self, "Exportado", "El presupuesto y el desglose AIU se han exportado a CSV.")

    def export_excel(self):
        """
        Exporta al archivo Excel usando la plantilla 'FORMATO EXPORTACION.xlsx' existente en la raíz.
        Llena cuatro hojas:
          - PRESUPUESTO: equivalente a la tabla de presupuesto con capítulos y subtotales
          - AIU: desglose de costos indirectos si está disponible
          - ANALISIS UNITARIOS: lista de análisis del presupuesto con totales actuales
          - INSUMOS: todos los recursos de los análisis incluidos, agrupados
        """
        try:
            # Seleccionar destino
            filePath, _ = QFileDialog.getSaveFileName(self, "Exportar Excel", "", "Excel (*.xlsx)")
            if not filePath:
                return
            template_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "FORMATO EXPORTACION.xlsx"))
            if not os.path.exists(template_path):
                QMessageBox.critical(self, "Plantilla no encontrada", f"No se encontró la plantilla: {template_path}")
                return

            # Cargar plantilla
            wb = load_workbook(template_path)

            # Obtener hojas (tolerante a nombres)
            def _get_ws(*candidates):
                for name in candidates:
                    if name in wb.sheetnames:
                        return wb[name]
                return wb.active

            ws_ppto = _get_ws("PRESUPUESTO", "Presupuesto")
            ws_aiu = _get_ws("AIU", "Aiu")
            ws_au = _get_ws("ANALISIS", "Analisis", "ANALISIS UNITARIOS", "Analisis Unitarios", "Análisis Unitarios")
            ws_ins = _get_ws("INSUMOS", "Insumos")

            # Helpers para celdas combinadas
            def _anchor(sheet, r, c):
                cell = sheet.cell(row=r, column=c)
                if isinstance(cell, MergedCell):
                    for cr in sheet.merged_cells.ranges:
                        if cr.min_row <= r <= cr.max_row and cr.min_col <= c <= cr.max_col:
                            return sheet.cell(row=cr.min_row, column=cr.min_col)
                return cell

            def _set(sheet, r, c, value):
                cell = sheet.cell(row=r, column=c)
                try:
                    cell.value = value
                    return cell
                except Exception:
                    # Si es una celda combinada, escribir en el ancla
                    base = _anchor(sheet, r, c)
                    try:
                        base.value = value
                    except Exception:
                        pass
                    return base

            # Conversor simple de números a letras en español (enteros hasta billones)
            def _num_a_letras_es(n: int) -> str:
                unidades = ["CERO","UNO","DOS","TRES","CUATRO","CINCO","SEIS","SIETE","OCHO","NUEVE",
                            "DIEZ","ONCE","DOCE","TRECE","CATORCE","QUINCE","DIECISEIS","DIECISIETE","DIECIOCHO","DIECINUEVE"]
                decenas = ["","DIEZ","VEINTE","TREINTA","CUARENTA","CINCUENTA","SESENTA","SETENTA","OCHENTA","NOVENTA"]
                centenas = ["","CIENTO","DOSCIENTOS","TRESCIENTOS","CUATROCIENTOS","QUINIENTOS","SEISCIENTOS","SETECIENTOS","OCHOCIENTOS","NOVECIENTOS"]
                def _tres(x: int) -> str:
                    if x == 0:
                        return ""
                    if x == 100:
                        return "CIEN"
                    c = x // 100
                    d = (x % 100) // 10
                    u = x % 10
                    resto = x % 100
                    res = []
                    if c:
                        res.append(centenas[c])
                    if resto <= 19:
                        if resto:
                            res.append(unidades[resto])
                        return " ".join([r for r in res if r])
                    if d == 2:
                        if u == 0:
                            res.append("VEINTE")
                        else:
                            res.append("VEINTI" + ("UN" if u == 1 else unidades[u].lower()))
                        return " ".join([r for r in res if r])
                    if d >= 3:
                        res.append(decenas[d])
                        if u:
                            res.append("Y")
                            res.append("UN" if u == 1 else unidades[u])
                        return " ".join([r for r in res if r])
                    return " ".join([r for r in res if r])
                if n == 0:
                    return "CERO"
                partes = []
                billones = n // 1_000_000_000_000
                n %= 1_000_000_000_000
                miles_mill = n // 1_000_000_000
                n %= 1_000_000_000
                millones = n // 1_000_000
                n %= 1_000_000
                miles = n // 1000
                resto = n % 1000
                if billones:
                    partes.append(_tres(billones))
                    partes.append("BILLONES" if billones > 1 else "BILLON")
                if miles_mill:
                    partes.append(_tres(miles_mill))
                    partes.append("MIL MILLONES")
                if millones:
                    partes.append(_tres(millones))
                    partes.append("MILLONES" if millones > 1 else "MILLON")
                if miles:
                    if miles == 1:
                        partes.append("MIL")
                    else:
                        partes.append(_tres(miles))
                        partes.append("MIL")
                if resto:
                    partes.append(_tres(resto))
                # Ajustes UN/UNO
                txt = " ".join([p for p in partes if p]).replace(" UNO MIL"," UN MIL").replace(" UNO MILLON"," UN MILLON").replace(" UNO MILLONES"," UN MILLONES")
                txt = " ".join(txt.split())
                return txt

            # -------- PRESUPUESTO --------
            # Buscar header donde dice ITEM
            header_row = None
            for r in range(1, ws_ppto.max_row + 1):
                v = ws_ppto.cell(row=r, column=1).value
                if isinstance(v, str) and v.strip().upper() == "ITEM":
                    header_row = r
                    break
            if header_row is None:
                header_row = 8
            data_row = header_row + 1

            # Guardar solo la información de merges (evitar copiar estilos para prevenir StyleProxy issues)
            def _row_style(sheet, r, cols=6):
                merges = []
                for cr in sheet.merged_cells.ranges:
                    if cr.min_row == r and cr.max_row == r and cr.min_col <= 6:
                        merges.append((cr.min_col, cr.max_col))
                return None, merges

            cap_style, cap_merges = _row_style(ws_ppto, data_row)
            item_style, item_merges = _row_style(ws_ppto, data_row + 1)
            sub_style, sub_merges = _row_style(ws_ppto, data_row + 2)

            def _apply_style(sheet, r, snap, merges):
                # Solo aplicar merges para imitar el formato de fila sin copiar estilos
                for mc, xc in merges:
                    try:
                        sheet.merge_cells(start_row=r, start_column=mc, end_row=r, end_column=xc)
                    except Exception:
                        pass

            # Limpiar bloque de datos actual: borrar filas de datos sin tocar el resumen
            # Buscar ancla del bloque de resumen por varias palabras clave
            def _find_anchor_row(sheet):
                # Priorizar el inicio real del bloque de totales
                priority = ["COSTOS INDIRECTOS"]
                fallback = ["VALOR COSTO", "COSTO DIRECTO", "TOTAL PRESUPUESTO", "FIRMA"]
                max_c = min(12, sheet.max_column or 12)
                # Búsqueda prioritaria
                for r in range(data_row, sheet.max_row + 1):
                    for c in range(1, max_c + 1):
                        v = sheet.cell(row=r, column=c).value
                        if isinstance(v, str) and any(k in v.upper() for k in priority):
                            return r
                # Fallback
                max_c = min(12, sheet.max_column or 12)
                for r in range(data_row, sheet.max_row + 1):
                    for c in range(1, max_c + 1):
                        v = sheet.cell(row=r, column=c).value
                        if isinstance(v, str) and any(k in v.upper() for k in fallback):
                            return r
                return None

            totals_anchor = _find_anchor_row(ws_ppto)
            if totals_anchor and totals_anchor > data_row:
                try:
                    ws_ppto.delete_rows(data_row, totals_anchor - data_row)
                except Exception:
                    pass

            # Construir capítulos desde la tabla
            chapters = []
            current = None
            for tr in range(self.table.rowCount()):
                it0 = self.table.item(tr, 0)
                if not it0:
                    continue
                role = it0.data(Qt.ItemDataRole.UserRole)
                if role == 'chapter':
                    if current:
                        chapters.append(current)
                    text = it0.text()
                    try:
                        num, name = text.split('.', 1)
                        num = num.strip()
                        name = name.strip()
                    except Exception:
                        num = str(len(chapters) + 1)
                        name = text
                    current = { 'num': num, 'name': name, 'items': [], 'subtotal': 0.0 }
                elif role == 'subtotal':
                    val_txt = self.table.item(tr, 5).text() if self.table.item(tr, 5) else "0"
                    try:
                        st = float(str(val_txt).replace('$','').replace(',',''))
                    except Exception:
                        st = 0.0
                    if current:
                        current['subtotal'] = st
                else:
                    if current is None:
                        continue
                    def _txt(c):
                        x = self.table.item(tr, c)
                        return x.text() if x else ""
                    def _num(c):
                        t = (_txt(c) or '').replace('$','').replace(' ','').replace(',','')
                        try:
                            return float(t)
                        except Exception:
                            return 0.0
                    current['items'].append({
                        'item': it0.text(),
                        'desc': _txt(1),
                        'und': (_txt(2) or '').upper(),
                        'qty': _num(3),
                        'cu': _num(4),
                        'ct': _num(5),
                    })
            if current:
                chapters.append(current)

            # Estilos para capítulo y subtotal (subrayado)
            chap_fill = XLFill(fill_type='solid', start_color='FFEFEFEF', end_color='FFEFEFEF')
            chap_font = XLFont(bold=True)
            side_thin = XLSide(style='thin', color='FFAAAAAA')
            side_med = XLSide(style='medium', color='FF000000')
            def _set_row_bottom_border(sheet, row_idx, left=2, right=6, border_side=None):
                for c in range(left, right + 1):
                    try:
                        cell = _set(sheet, row_idx, c, sheet.cell(row=row_idx, column=c).value)
                        cell.border = XLBorder(left=side_thin, right=side_thin, top=side_thin, bottom=(border_side or side_med))
                    except Exception:
                        pass

            rptr = data_row
            for ch in chapters:
                ws_ppto.insert_rows(rptr, 1)
                _apply_style(ws_ppto, rptr, cap_style, cap_merges)
                try:
                    ws_ppto.merge_cells(start_row=rptr, start_column=2, end_row=rptr, end_column=6)
                except Exception:
                    pass
                _set(ws_ppto, rptr, 1, ch['num'])
                _set(ws_ppto, rptr, 2, f"CAP {ch['num']}  {ch['name']}")
                # Estilo de capítulo: fondo y subrayado (borde inferior)
                try:
                    for c in range(1, 7):
                        cell = ws_ppto.cell(row=rptr, column=c)
                        cell.fill = chap_fill
                        cell.font = chap_font
                        if c == 2:
                            cell.alignment = XLAlign(horizontal='left', vertical='center')
                        cell.border = XLBorder(left=side_thin, right=side_thin, top=side_thin, bottom=side_med)
                    ws_ppto.row_dimensions[rptr].height = 18
                except Exception:
                    pass
                rptr += 1
                for it in ch['items']:
                    ws_ppto.insert_rows(rptr, 1)
                    _apply_style(ws_ppto, rptr, item_style, item_merges)
                    _set(ws_ppto, rptr, 1, it['item'])
                    _set(ws_ppto, rptr, 2, it['desc'])
                    _set(ws_ppto, rptr, 3, it['und'])
                    _set(ws_ppto, rptr, 4, it['qty'])
                    _set(ws_ppto, rptr, 5, it['cu'])
                    _set(ws_ppto, rptr, 6, it['ct'])
                    # bordes suaves laterales
                    try:
                        for c in range(1,7):
                            cell = ws_ppto.cell(row=rptr, column=c)
                            cell.border = XLBorder(left=side_thin, right=side_thin)
                        ws_ppto.row_dimensions[rptr].height = 18
                    except Exception:
                        pass
                    rptr += 1
                ws_ppto.insert_rows(rptr, 1)
                _apply_style(ws_ppto, rptr, sub_style, sub_merges)
                try:
                    ws_ppto.merge_cells(start_row=rptr, start_column=2, end_row=rptr, end_column=5)
                except Exception:
                    pass
                _set(ws_ppto, rptr, 2, f"SUBTOTAL CAP {ch['num']}  {ch['name']}: ")
                _set(ws_ppto, rptr, 6, ch['subtotal'])
                # Estilo de subtotal: negrita y borde superior grueso
                try:
                    for c in range(1,7):
                        cell = ws_ppto.cell(row=rptr, column=c)
                        cell.font = chap_font
                        cell.border = XLBorder(top=side_med, left=side_thin, right=side_thin)
                        if c == 2:
                            cell.alignment = XLAlign(horizontal='right', vertical='center')
                        if c == 6:
                            cell.number_format = '#,##0.00'
                    ws_ppto.row_dimensions[rptr].height = 18
                except Exception:
                    pass
                rptr += 1
                # Separación entre tablas de capítulos
                rptr += 1

            # Encontrar y mover la fila "VALOR COSTOS DIRECTOS" a la posición correcta
            try:
                # Buscar la fila que contiene "VALOR COSTOS DIRECTOS"
                vcd_row = None
                for row in range(1, ws_ppto.max_row + 1):
                    for col in range(1, ws_ppto.max_column + 1):
                        try:
                            cell_value = ws_ppto.cell(row=row, column=col).value
                            if cell_value and isinstance(cell_value, str):
                                if 'VALOR COSTOS DIRECTOS' in cell_value.upper():
                                    if 'COSTOS INDIRECTOS' not in cell_value.upper():  # No tocar "COSTOS INDIRECTOS"
                                        vcd_row = row
                                        break
                        except Exception:
                            continue
                    if vcd_row:
                        break
                
                # Si encontramos la fila, copiar su contenido a la posición correcta y limpiar la original
                if vcd_row and vcd_row != rptr:
                    # Copiar todo el contenido de la fila original a la nueva posición
                    for col in range(1, ws_ppto.max_column + 1):
                        try:
                            original_cell = ws_ppto.cell(row=vcd_row, column=col)
                            new_cell = ws_ppto.cell(row=rptr, column=col)
                            new_cell.value = original_cell.value
                            if original_cell.has_style:
                                new_cell.font = original_cell.font
                                new_cell.fill = original_cell.fill
                                new_cell.border = original_cell.border
                                new_cell.alignment = original_cell.alignment
                                new_cell.number_format = original_cell.number_format
                        except Exception:
                            continue
                    
                    # Limpiar la fila original
                    for col in range(1, ws_ppto.max_column + 1):
                        try:
                            ws_ppto.cell(row=vcd_row, column=col).value = None
                        except Exception:
                            continue
                    
                    # Actualizar el valor con el costo directo correcto
                    _set(ws_ppto, rptr, 6, direct_cost)
                    
                    # Asegurar altura de fila
                    try:
                        ws_ppto.row_dimensions[rptr].height = 18
                    except Exception:
                        pass
                else:
                    # Si no se encontró, crear desde cero
                    _set(ws_ppto, rptr, 2, 'VALOR COSTOS DIRECTOS')
                    _set(ws_ppto, rptr, 6, direct_cost)
                    try:
                        ws_ppto.merge_cells(start_row=rptr, start_column=2, end_row=rptr, end_column=5)
                    except Exception:
                        pass
                    # Estilo
                    try:
                        for c in range(2,7):
                            cell = ws_ppto.cell(row=rptr, column=c)
                            cell.font = XLFont(bold=True)
                            cell.alignment = XLAlign(horizontal='right' if c==2 else 'center', vertical='center')
                            if c == 6:
                                cell.number_format = '#,##0.00'
                        ws_ppto.row_dimensions[rptr].height = 18
                    except Exception:
                        pass
                
                # Asegurar que "COSTOS INDIRECTOS" esté en la siguiente fila
                summary_anchor = rptr + 1
                # Buscar si ya existe "COSTOS INDIRECTOS" en la fila siguiente
                costos_exists = False
                try:
                    for col in range(1, ws_ppto.max_column + 1):
                        cell_value = ws_ppto.cell(row=summary_anchor, column=col).value
                        if cell_value and isinstance(cell_value, str) and 'COSTOS INDIRECTOS' in cell_value.upper():
                            costos_exists = True
                            break
                except Exception:
                    pass
                
                if not costos_exists:
                    _set(ws_ppto, summary_anchor, 2, 'COSTOS INDIRECTOS')
                    try:
                        ws_ppto.merge_cells(start_row=summary_anchor, start_column=2, end_row=summary_anchor, end_column=6)
                    except Exception:
                        pass
                    try:
                        cell = ws_ppto.cell(row=summary_anchor, column=2)
                        cell.font = XLFont(bold=True)
                        cell.fill = XLFill(fill_type='solid', start_color='FFEFEFEF', end_color='FFEFEFEF')
                        cell.alignment = XLAlign(horizontal='left', vertical='center')
                        ws_ppto.row_dimensions[summary_anchor].height = 18
                    except Exception:
                        pass
                        
            except Exception:
                pass

            # Escribir totales buscando por sinónimos si es necesario
            import unicodedata
            def _norm(s: str) -> str:
                s = unicodedata.normalize('NFKD', s)
                s = ''.join(ch for ch in s if not unicodedata.combining(ch))
                return s.upper()

            def _write_by_keywords(sheet, labels: list[str], value, col=6):
                labels_n = [_norm(x) for x in labels]
                for rr in range(1, sheet.max_row + 1):
                    for cc in range(1, min(12, sheet.max_column or 12) + 1):
                        v = sheet.cell(row=rr, column=cc).value
                        if not isinstance(v, str):
                            continue
                        up = _norm(v)
                        if any(lbl in up for lbl in labels_n):
                            cell = _set(sheet, rr, col, value)
                            return True
                return False

            def _write_all_by_keywords(sheet, labels: list[str], value, col=6):
                """Escribe en TODAS las coincidencias de etiquetas en la columna indicada."""
                labels_n = [_norm(x) for x in labels]
                wrote = False
                for rr in range(1, sheet.max_row + 1):
                    for cc in range(1, min(12, sheet.max_column or 12) + 1):
                        v = sheet.cell(row=rr, column=cc).value
                        if not isinstance(v, str):
                            continue
                        up = _norm(v)
                        if any(lbl in up for lbl in labels_n):
                            cell = _set(sheet, rr, col, value)
                            wrote = True
                return wrote

            def _find_label_rows(sheet, labels: list[str]):
                rows = []
                labels_n = [_norm(x) for x in labels]
                for rr in range(1, sheet.max_row + 1):
                    for cc in range(1, min(12, sheet.max_column or 12) + 1):
                        v = sheet.cell(row=rr, column=cc).value
                        if isinstance(v, str) and any(lbl in _norm(v) for lbl in labels_n):
                            rows.append(rr)
                            break
                return rows

            def _write_last_by_keywords(sheet, labels: list[str], value, col=6):
                rows = _find_label_rows(sheet, labels)
                if not rows:
                    return False
                target_row = max(rows)
                _set(sheet, target_row, col, value)
                return True

            # Escribir líneas de resumen con % en col 5 y valor en col 6
            def _write_summary_line(sheet, labels: list[str], pct_value: float, amount_value: float, pct_col=5, val_col=6, bold=False):
                labels_n = [_norm(x) for x in labels]
                for rr in range(1, sheet.max_row + 1):
                    for cc in range(1, min(12, sheet.max_column or 12) + 1):
                        v = sheet.cell(row=rr, column=cc).value
                        if not isinstance(v, str):
                            continue
                        up = _norm(v)
                        if any(lbl in up for lbl in labels_n):
                            pct_cell = _set(sheet, rr, pct_col, None)
                            val_cell = _set(sheet, rr, val_col, None)
                            # Porcentajes se escriben como 0-1 con formato 0.00%
                            try:
                                pct_cell.value = (pct_value or 0.0) / 100.0
                                pct_cell.number_format = '0.00%'
                            except Exception:
                                pct_cell.value = pct_value
                            try:
                                val_cell.value = amount_value or 0.0
                                val_cell.number_format = '#,##0.00'
                            except Exception:
                                val_cell.value = amount_value
                            if bold:
                                try:
                                    bf = XLFont(bold=True)
                                    pct_cell.font = bf
                                    val_cell.font = bf
                                except Exception:
                                    pass
                            return True
                return False

            direct_cost = getattr(self, 'direct_cost_total', 0.0)
            admin_total = getattr(self, 'admin_cost_total', 0.0)
            # VALOR COSTOS DIRECTOS ya se escribió arriba después del último capítulo

            # Desglose en resumen del Presupuesto (ADMIN/IMPREV/UTIL/TOTAL AIU/IVA)
            bd = getattr(self, 'aiu_breakdown', None)
            admin_val = bd.get('admin', 0.0) if bd else 0.0
            imprev_val = bd.get('imprev', 0.0) if bd else 0.0
            util_val = bd.get('util', 0.0) if bd else 0.0
            iva_val = bd.get('iva', 0.0) if bd else 0.0
            admin_pct = (admin_val / direct_cost * 100.0) if direct_cost > 0 else 0.0
            imprev_pct = (bd.get('imprev_pct', 0.0) if bd else 0.0)
            util_pct = (bd.get('util_pct', 0.0) if bd else 0.0)
            iva_pct = (bd.get('iva_pct', 19.0) if bd else 0.0)
            aiu_sin_iva_val = admin_val + imprev_val + util_val
            aiu_total_val = aiu_sin_iva_val + iva_val
            aiu_sin_iva_pct = (aiu_sin_iva_val / direct_cost * 100.0) if direct_cost > 0 else (admin_pct + imprev_pct + util_pct)
            aiu_pct_total = (aiu_total_val / direct_cost * 100.0) if direct_cost > 0 else (aiu_sin_iva_pct + iva_pct)

            rows_to_uniform = []
            for lbls, pct, val, bold in [
                (['ADMINISTRACION', 'ADMINISTRACIÓN'], admin_pct, admin_val, False),
                (['IMPREVISTOS'], imprev_pct, imprev_val, False),
                (['UTILIDAD'], util_pct, util_val, False),
                # TOTAL AIU (solo AIU sin IVA)
                (['TOTAL AIU'], aiu_sin_iva_pct, aiu_sin_iva_val, True),
                (['COSTOS INDIRECTOS'], aiu_pct_total, aiu_total_val, True),
                (['IVA SOBRE LA UTILIDAD', 'IVA UTILIDAD'], iva_pct, iva_val, False),
            ]:
                # registrar fila usada por cada escritura
                labels_n = [_norm(x) for x in lbls]
                used_row = None
                for rr in range(1, ws_ppto.max_row + 1):
                    found = False
                    for cc in range(1, min(12, ws_ppto.max_column or 12) + 1):
                        v = ws_ppto.cell(row=rr, column=cc).value
                        if isinstance(v, str) and any(l in _norm(v) for l in labels_n):
                            used_row = rr
                            found = True
                            break
                    if found:
                        break
                _write_summary_line(ws_ppto, lbls, pct, val, bold=bold)
                if used_row:
                    rows_to_uniform.append(used_row)

            # Uniformar altura de filas en el bloque de resumen (incluye fila 23 solicitada)
            try:
                for rr in rows_to_uniform + [23]:
                    ws_ppto.row_dimensions[rr].height = 18
            except Exception:
                pass

            _write_last_by_keywords(ws_ppto, ['VALOR TOTAL PRESUPUESTO', 'VALOR COSTO TOTAL', 'TOTAL PRESUPUESTO'], direct_cost + admin_total)

            # Reemplazar valor por fórmula que sume las filas del propio resumen (Directo + Total AIU + IVA)
            try:
                # Buscar la tabla de resumen correcta (la última sección que tenga la etiqueta "COSTOS INDIRECTOS")
                costos_rows = _find_label_rows(ws_ppto, ['COSTOS INDIRECTOS'])
                vcd_rows = _find_label_rows(ws_ppto, ['VALOR COSTOS DIRECTOS', 'VALOR COSTO DIRECTO', 'COSTO DIRECTO'])
                total_aiu_rows = _find_label_rows(ws_ppto, ['TOTAL AIU'])
                iva_rows = _find_label_rows(ws_ppto, ['IVA SOBRE LA UTILIDAD', 'IVA UTILIDAD'])
                total_rows = _find_label_rows(ws_ppto, ['VALOR TOTAL PRESUPUESTO', 'VALOR COSTO TOTAL', 'TOTAL PRESUPUESTO'])
                if costos_rows and vcd_rows and total_aiu_rows and iva_rows and total_rows:
                    base = max(costos_rows)  # ancla del bloque final
                    # Seleccionar filas posteriores a la ancla para evitar líneas intermedias
                    vcd_candidates = [r for r in vcd_rows if r > base]
                    taiu_candidates = [r for r in total_aiu_rows if r > base]
                    iva_candidates = [r for r in iva_rows if r > base]
                    total_candidates = [r for r in total_rows if r > base]
                    if vcd_candidates and taiu_candidates and iva_candidates and total_candidates:
                        vcd_r = min(vcd_candidates)
                        taiu_r = min(taiu_candidates)
                        iva_r = min(iva_candidates)
                        tot_r = min(total_candidates)
                        # Fórmula en la columna 6 (F) del total
                        cell = _set(ws_ppto, tot_r, 6, f"=F{vcd_r}+F{taiu_r}+F{iva_r}")
                        try:
                            cell.number_format = '#,##0.00'
                        except Exception:
                            pass
            except Exception:
                pass

            # Escribir valor en letras: siempre en la fila inmediatamente debajo de
            # la etiqueta "VALOR  TOTAL PRESUPUESTO"
            try:
                vt_rows = _find_label_rows(ws_ppto, ['VALOR  TOTAL PRESUPUESTO', 'VALOR TOTAL PRESUPUESTO'])
                if vt_rows:
                    target_row = max(vt_rows) + 1
                    total_entero = int(round(direct_cost + admin_total))
                    letras = _num_a_letras_es(total_entero)
                    target_c = 2
                    _set(ws_ppto, target_row, target_c, letras)
                    for c in range(target_c, min(target_c + 4, ws_ppto.max_column + 1)):
                        try:
                            ws_ppto.cell(row=target_row, column=c).alignment = XLAlign(horizontal='left', vertical='center', wrap_text=True)
                        except Exception:
                            pass
            except Exception:
                pass

            # -------- AIU --------
            if bd:
                _write_all_by_keywords(ws_aiu, ['VALOR COSTO DIRECTO', 'COSTO DIRECTO'], direct_cost)
                _write_summary_line(ws_aiu, ['ADMINISTRACION', 'ADMINISTRACIÓN'], admin_pct, admin_val)
                _write_summary_line(ws_aiu, ['IMPREVISTOS'], imprev_pct, imprev_val)
                _write_summary_line(ws_aiu, ['UTILIDAD'], util_pct, util_val)
                _write_summary_line(ws_aiu, ['IVA SOBRE LA UTILIDAD', 'IVA UTILIDAD'], iva_pct, iva_val)
                _write_summary_line(ws_aiu, ['TOTAL AIU', 'COSTOS INDIRECTOS'], aiu_pct_total, aiu_total_val, bold=True)
                _write_all_by_keywords(ws_aiu, ['VALOR COSTO TOTAL', 'TOTAL PRESUPUESTO'], direct_cost + bd.get('total_aiu', 0.0))

            # -------- ANALISIS UNITARIOS --------
            # Limpiar cuerpo debajo de header (si existe) y construir una tabla por cada análisis
            header_row_au = 1
            for r in range(1, ws_au.max_row + 1):
                v = ws_au.cell(row=r, column=1).value
                if isinstance(v, str) and v.strip().upper() in ("CODIGO", "CÓDIGO", "ANALISIS", "ANÁLISIS"):
                    header_row_au = r
                    break
            try:
                ws_au.delete_rows(header_row_au + 1, max(ws_au.max_row - header_row_au, 0))
            except Exception:
                pass

            # Extraer análisis únicos desde la tabla de presupuesto
            seen = set()
            analyses = []
            for r in range(self.table.rowCount()):
                it0 = self.table.item(r, 0)
                if not it0:
                    continue
                role = it0.data(Qt.ItemDataRole.UserRole)
                if role in (None, '', 'chapter', 'subtotal'):
                    continue
                code = role
                if code in seen:
                    continue
                seen.add(code)
                desc = self.table.item(r, 1).text() if self.table.item(r, 1) else ''
                und = self.table.item(r, 2).text() if self.table.item(r, 2) else ''
                analyses.append((code, desc, und))

            # Estilos básicos (sin StyleProxy) y ajuste de columnas
            title_font = XLFont(bold=True, size=12)
            header_font = XLFont(bold=True)
            center = XLAlign(horizontal='center', vertical='center', wrap_text=True)
            left = XLAlign(horizontal='left', vertical='center', wrap_text=True)
            num_fmt = '#,##0.00'
            header_fill = XLFill('solid', start_color='FFEFEFEF', end_color='FFEFEFEF')
            thin = XLSide(style='thin', color='FF999999')
            border = XLBorder(left=thin, right=thin, top=thin, bottom=thin)

            try:
                ws_au.column_dimensions['A'].width = 16
                ws_au.column_dimensions['B'].width = 60
                ws_au.column_dimensions['C'].width = 10
                ws_au.column_dimensions['D'].width = 12
                ws_au.column_dimensions['E'].width = 14
                ws_au.column_dimensions['F'].width = 16
            except Exception:
                pass

            rptr = header_row_au + 1
            from models.database import SessionLocal as _Sess
            _session = _Sess()
            try:
                for code, desc, und in analyses:
                    # Título del análisis (merge A..F)
                    _set(ws_au, rptr, 1, f"{code} - {desc}")
                    try:
                        ws_au.merge_cells(start_row=rptr, start_column=1, end_row=rptr, end_column=6)
                    except Exception:
                        pass
                    for c in range(1, 7):
                        cell = ws_au.cell(row=rptr, column=c)
                        cell.font = title_font
                        cell.alignment = left
                    rptr += 1

                    # Header de la tabla de recursos
                    headers = ["CÓDIGO", "DESCRIPCIÓN", "UND", "CANT.", "VR. UNIT", "VR. TOTAL"]
                    for c, h in enumerate(headers, start=1):
                        cell = ws_au.cell(row=rptr, column=c, value=h)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = center
                        cell.border = border
                    rptr += 1

                    # Filas de recursos
                    total_parcial = 0.0
                    q = _session.query(AnalisisUnitarioRecurso).filter_by(codigo_analisis=code).all()
                    if not q:
                        # Fila vacía para indicar sin datos
                        ws_au.cell(row=rptr, column=1, value="(sin recursos)").alignment = left
                        rptr += 1
                    else:
                        for rec in q:
                            cant = float(rec.cantidad_recurso or 0)
                            desper = float(rec.desper or 0)
                            vr_unit = float(rec.vr_unitario or 0)
                            vr_parc = cant * (1 + desper) * vr_unit
                            total_parcial += vr_parc
                            values = [
                                rec.codigo_recurso,
                                rec.descripcion_recurso,
                                rec.unidad_recurso,
                                cant,
                                vr_unit,
                                vr_parc
                            ]
                            for c, val in enumerate(values, start=1):
                                cell = ws_au.cell(row=rptr, column=c, value=val)
                                if c in (1,2):
                                    cell.alignment = left
                                else:
                                    cell.alignment = center
                                if c in (4,5,6):
                                    cell.number_format = num_fmt
                                cell.border = border
                            rptr += 1

                    # Subtotal del análisis
                    try:
                        ws_au.merge_cells(start_row=rptr, start_column=1, end_row=rptr, end_column=5)
                    except Exception:
                        pass
                    st_cell = ws_au.cell(row=rptr, column=1, value=f"SUBTOTAL {code}:")
                    st_cell.font = header_font
                    st_cell.alignment = left
                    val_cell = ws_au.cell(row=rptr, column=6, value=total_parcial)
                    val_cell.number_format = num_fmt
                    val_cell.font = header_font
                    rptr += 2  # una fila en blanco entre tablas
            finally:
                _session.close()

            # -------- INSUMOS --------
            # Limpiar cuerpo debajo del header (buscar 'CODIGO RECURSO')
            header_row_ins = 1
            for r in range(1, ws_ins.max_row + 1):
                v = ws_ins.cell(row=r, column=1).value
                if isinstance(v, str) and 'CODIGO' in v.upper():
                    header_row_ins = r
                    break
            ws_ins.delete_rows(header_row_ins + 1, ws_ins.max_row - header_row_ins)

            # Consultar BD para traer recursos de cada análisis incluido
            from models.database import SessionLocal
            session = SessionLocal()
            try:
                rptr = header_row_ins + 1
                for code, _, _ in analyses:
                    # insertar encabezado
                    _set(ws_ins, rptr, 1, f"=== {code} ===")
                    rptr += 1
                    q = session.query(AnalisisUnitarioRecurso).filter_by(codigo_analisis=code).all()
                    for rec in q:
                        _set(ws_ins, rptr, 1, rec.codigo_recurso)
                        _set(ws_ins, rptr, 2, rec.descripcion_recurso)
                        _set(ws_ins, rptr, 3, rec.unidad_recurso)
                        _set(ws_ins, rptr, 4, rec.cantidad_recurso)
                        _set(ws_ins, rptr, 5, rec.desper)
                        _set(ws_ins, rptr, 6, rec.vr_unitario)
                        _set(ws_ins, rptr, 7, rec.vr_parcial)
                        rptr += 1
            finally:
                session.close()

            # Guardar salida
            wb.save(filePath)
            QMessageBox.information(self, "Exportado", "Se generó el archivo Excel con el formato de la plantilla.")
        except Exception as e:
            QMessageBox.critical(self, "Error al exportar", f"{e}")

    def export_excel_from_scratch(self):
        """Exporta el presupuesto a un archivo Excel construido completamente desde cero"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Seleccionar destino
            filePath, _ = QFileDialog.getSaveFileName(self, "Exportar Excel", "Presupuesto_Exportado.xlsx", "Excel (*.xlsx)")
            if not filePath:
                return
            
            # Crear un nuevo workbook
            wb = Workbook()
            
            # =====================
            # HOJA PRESUPUESTO
            # =====================
            ws = wb.active
            ws.title = "PRESUPUESTO"
            
            # Configurar anchos de columna
            ws.column_dimensions['A'].width = 8    # ITEM
            ws.column_dimensions['B'].width = 50   # DESCRIPCION  
            ws.column_dimensions['C'].width = 8    # UND
            ws.column_dimensions['D'].width = 12   # CANT
            ws.column_dimensions['E'].width = 15   # VR. UNIT
            ws.column_dimensions['F'].width = 15   # VR.TOTAL
            
            # Estilos reutilizables
            header_font = Font(bold=True, size=12)
            chapter_font = Font(bold=True, size=11)
            subtotal_font = Font(bold=True, size=10)
            normal_font = Font(size=10)
            
            header_fill = PatternFill(fill_type='solid', start_color='FFCCCCCC')
            chapter_fill = PatternFill(fill_type='solid', start_color='FFEFEFEF')
            
            center_align = Alignment(horizontal='center', vertical='center')
            left_align = Alignment(horizontal='left', vertical='center')
            right_align = Alignment(horizontal='right', vertical='center')
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            thick_bottom = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thick')
            )
            
            thick_top = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thick'),
                bottom=Side(style='thin')
            )
            
            # Header del documento
            ws.merge_cells('A1:C1')
            ws['A1'] = "NOMBRE DE LA EMPRESA"
            ws['A1'].font = header_font
            ws['A1'].alignment = center_align
            ws['A1'].border = thin_border
            
            ws.merge_cells('D1:F1')
            ws['D1'] = "LOGO DE LA EMPRESA"
            ws['D1'].font = header_font
            ws['D1'].alignment = center_align
            ws['D1'].border = thin_border
            
            # Ajustar altura de la primera fila
            ws.row_dimensions[1].height = 25
            
            # Título principal
            ws.merge_cells('A3:F3')
            ws['A3'] = "PRESUPUESTO DE OBRA"
            ws['A3'].font = Font(bold=True, size=14)
            ws['A3'].alignment = center_align
            ws['A3'].fill = header_fill
            ws['A3'].border = thin_border
            
            # Información de la obra
            ws['A4'] = "Obra:"
            ws['A4'].font = header_font
            ws.merge_cells('B4:D4')
            ws['B4'] = "ESCRIBA AQUÍ EL NOMBRE DE LA OBRA"
            ws['B4'].border = thin_border
            
            ws['E4'] = "FECHA:"
            ws['E4'].font = header_font
            ws['F4'] = "19-sept-25"
            ws['F4'].border = thin_border
            
            ws.merge_cells('E5:F5')
            ws['E5'] = "QUIEN ELABORÓ:"
            ws['E5'].font = header_font
            ws['E5'].border = thin_border
            
            # Headers de la tabla
            row = 7
            headers = ["ITEM", "DESCRIPCION", "UND", "CANT.", "VR. UNIT", "VR.TOTAL"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
            
            # Obtener datos del presupuesto usando UserRole
            from PyQt6.QtCore import Qt
            chapters = []
            current = None
            
            for tr in range(self.table.rowCount()):
                it0 = self.table.item(tr, 0)
                if not it0:
                    continue
                    
                # Usar UserRole para determinar el tipo de fila
                role = it0.data(Qt.ItemDataRole.UserRole)
                
                # Detectar si es capítulo
                if role == 'chapter':
                    if current:
                        chapters.append(current)
                    desc1 = self.table.item(tr, 1)
                    current = {
                        'title': it0.text(),
                        'desc': desc1.text() if desc1 else '',
                        'items': [],
                        'subtotal': 0.0
                    }
                # Detectar si es subtotal
                elif role == 'subtotal':
                    val_item = self.table.item(tr, 5)  # Columna VR.TOTAL
                    val_txt = val_item.text() if val_item else '0'
                    try:
                        st = float(str(val_txt).replace('$','').replace(',','').replace(' ',''))
                    except Exception:
                        st = 0.0
                    if current:
                        current['subtotal'] = st
                # Es un análisis
                else:
                    if current is None:
                        continue
                    def _txt(c):
                        x = self.table.item(tr, c)
                        return x.text() if x else ""
                    def _num(c):
                        t = (_txt(c) or '').replace('$','').replace(',','').replace(' ','')
                        try:
                            return float(t)
                        except Exception:
                            return 0.0
                    current['items'].append({
                        'item': it0.text(),
                        'desc': _txt(1),
                        'und': (_txt(2) or '').upper(),
                        'qty': _num(3),
                        'cu': _num(4),
                        'ct': _num(5),
                    })
            if current:
                chapters.append(current)
            
            # Debug: verificar si hay datos
            if not chapters:
                QMessageBox.warning(self, "Sin datos", "No se encontraron capítulos en el presupuesto. Asegúrate de que el presupuesto tenga datos.")
                return
                
            # Escribir cada capítulo como una subtabla
            current_row = 8
            for ch in chapters:
                # Header del capítulo
                ws.merge_cells(f'A{current_row}:F{current_row}')
                cell = ws.cell(row=current_row, column=1, value=f"{ch['title']} {ch['desc']}")
                cell.font = chapter_font
                cell.fill = chapter_fill
                cell.alignment = left_align
                cell.border = thick_bottom
                ws.row_dimensions[current_row].height = 20
                current_row += 1
                
                # Items del capítulo
                for item in ch['items']:
                    ws.cell(row=current_row, column=1, value=item['item']).font = normal_font
                    ws.cell(row=current_row, column=2, value=item['desc']).font = normal_font
                    ws.cell(row=current_row, column=3, value=item['und']).font = normal_font
                    ws.cell(row=current_row, column=4, value=item['qty']).font = normal_font
                    ws.cell(row=current_row, column=5, value=item['cu']).font = normal_font
                    ws.cell(row=current_row, column=6, value=item['ct']).font = normal_font
                    
                    # Alineaciones y formatos
                    ws.cell(row=current_row, column=1).alignment = center_align
                    ws.cell(row=current_row, column=2).alignment = left_align
                    ws.cell(row=current_row, column=3).alignment = center_align
                    ws.cell(row=current_row, column=4).alignment = center_align
                    ws.cell(row=current_row, column=5).alignment = right_align
                    ws.cell(row=current_row, column=6).alignment = right_align
                    
                    # Formatos numéricos
                    ws.cell(row=current_row, column=4).number_format = '#,##0.00'
                    ws.cell(row=current_row, column=5).number_format = '#,##0.00'
                    ws.cell(row=current_row, column=6).number_format = '#,##0.00'
                    
                    # Bordes laterales
                    for col in range(1, 7):
                        ws.cell(row=current_row, column=col).border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin')
                        )
                    
                    ws.row_dimensions[current_row].height = 18
                    current_row += 1
                
                # Subtotal del capítulo
                ws.merge_cells(f'A{current_row}:E{current_row}')
                ws.cell(row=current_row, column=1, value=f"SUBTOTAL {ch['title']}:")
                ws.cell(row=current_row, column=1).font = subtotal_font
                ws.cell(row=current_row, column=1).alignment = right_align
                
                ws.cell(row=current_row, column=6, value=ch['subtotal'])
                ws.cell(row=current_row, column=6).font = subtotal_font
                ws.cell(row=current_row, column=6).alignment = right_align
                ws.cell(row=current_row, column=6).number_format = '#,##0.00'
                
                # Borde superior grueso para el subtotal
                for col in range(1, 7):
                    ws.cell(row=current_row, column=col).border = thick_top
                
                ws.row_dimensions[current_row].height = 18
                current_row += 2  # Espacio entre tablas
            
            # Tabla de resumen
            direct_cost = sum(ch['subtotal'] for ch in chapters)
            
            # VALOR COSTOS DIRECTOS
            ws.merge_cells(f'A{current_row}:E{current_row}')
            ws.cell(row=current_row, column=1, value="VALOR COSTOS DIRECTOS")
            ws.cell(row=current_row, column=1).font = subtotal_font
            ws.cell(row=current_row, column=1).alignment = right_align
            ws.cell(row=current_row, column=1).fill = chapter_fill
            
            ws.cell(row=current_row, column=6, value=direct_cost)
            ws.cell(row=current_row, column=6).font = subtotal_font
            ws.cell(row=current_row, column=6).alignment = right_align
            ws.cell(row=current_row, column=6).number_format = '#,##0.00'
            ws.cell(row=current_row, column=6).fill = chapter_fill
            
            for col in range(1, 7):
                ws.cell(row=current_row, column=col).border = thin_border
            
            ws.row_dimensions[current_row].height = 18
            current_row += 2
            
            # COSTOS INDIRECTOS header
            ws.merge_cells(f'A{current_row}:F{current_row}')
            ws.cell(row=current_row, column=1, value="COSTOS INDIRECTOS")
            ws.cell(row=current_row, column=1).font = subtotal_font
            ws.cell(row=current_row, column=1).alignment = left_align
            ws.cell(row=current_row, column=1).fill = chapter_fill
            
            # Aplicar bordes a toda la fila del header
            for col in range(1, 7):
                ws.cell(row=current_row, column=col).border = thin_border
                ws.cell(row=current_row, column=col).fill = chapter_fill
            
            ws.row_dimensions[current_row].height = 18
            current_row += 1
            
            # Obtener valores AIU
            bd = getattr(self, 'aiu_breakdown', None)
            admin_val = bd.get('admin', 0.0) if bd else 0.0
            imprev_val = bd.get('imprev', 0.0) if bd else 0.0
            util_val = bd.get('util', 0.0) if bd else 0.0
            iva_val = bd.get('iva', 0.0) if bd else 0.0
            
            # Calcular porcentajes
            admin_pct = (admin_val / direct_cost * 100) if direct_cost > 0 else 0
            imprev_pct = (imprev_val / direct_cost * 100) if direct_cost > 0 else 0
            util_pct = (util_val / direct_cost * 100) if direct_cost > 0 else 0
            total_aiu_no_iva = admin_val + imprev_val + util_val
            total_aiu_pct = (total_aiu_no_iva / direct_cost * 100) if direct_cost > 0 else 0
            iva_pct = (iva_val / (direct_cost + total_aiu_no_iva) * 100) if (direct_cost + total_aiu_no_iva) > 0 else 0
            
            # Líneas del desglose AIU
            aiu_items = [
                ("VALOR COSTOS DIRECTOS", direct_cost, "", True),
                ("ADMINISTRACIÓN", admin_val, f"{admin_pct:.2f}%", False),
                ("IMPREVISTOS", imprev_val, f"{imprev_pct:.2f}%", False),
                ("UTILIDAD", util_val, f"{util_pct:.2f}%", False),
                ("TOTAL AIU", total_aiu_no_iva, f"{total_aiu_pct:.2f}%", True),
                ("IVA SOBRE LA UTILIDAD", iva_val, f"{iva_pct:.2f}%", False),
                ("VALOR TOTAL PRESUPUESTO", direct_cost + admin_val + imprev_val + util_val + iva_val, "", True)
            ]
            
            for i, (label, value, percentage, is_bold) in enumerate(aiu_items):
                # Columna del concepto (D)
                ws.cell(row=current_row, column=4, value=label)
                ws.cell(row=current_row, column=4).font = subtotal_font if is_bold else normal_font
                ws.cell(row=current_row, column=4).alignment = right_align
                ws.cell(row=current_row, column=4).border = thin_border
                
                # Columna del porcentaje (E)
                if percentage:
                    ws.cell(row=current_row, column=5, value=percentage)
                    ws.cell(row=current_row, column=5).font = normal_font
                    ws.cell(row=current_row, column=5).alignment = center_align
                ws.cell(row=current_row, column=5).border = thin_border
                
                # Columna del valor (F)
                ws.cell(row=current_row, column=6, value=value)
                ws.cell(row=current_row, column=6).font = subtotal_font if is_bold else normal_font
                ws.cell(row=current_row, column=6).alignment = right_align
                ws.cell(row=current_row, column=6).number_format = '#,##0.00'
                ws.cell(row=current_row, column=6).border = thin_border
                
                # Bordes para columnas A, B, C (vacías pero con borde)
                for col in range(1, 4):
                    ws.cell(row=current_row, column=col).border = thin_border
                
                # Formato especial para filas importantes (fondo gris en toda la fila)
                if is_bold:
                    for col in range(1, 7):  # De A a F (columnas 1-6)
                        ws.cell(row=current_row, column=col).fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
                
                ws.row_dimensions[current_row].height = 18
                current_row += 1
            
            # Total en letras
            current_row += 1
            ws.merge_cells(f'A{current_row}:F{current_row}')
            total_value = direct_cost + admin_val + imprev_val + util_val + iva_val
            letters_text = self._num_a_letras_es(total_value)
            ws.cell(row=current_row, column=1, value=f"VALOR TOTAL PRESUPUESTO: {letters_text}")
            ws.cell(row=current_row, column=1).font = normal_font
            ws.cell(row=current_row, column=1).alignment = left_align
            ws.cell(row=current_row, column=1).border = thin_border
            ws.row_dimensions[current_row].height = 25
            current_row += 3
            
            # Plazo de entrega
            ws.merge_cells(f'C{current_row}:D{current_row}')
            ws.cell(row=current_row, column=3, value="PLAZO DE ENTREGA: (DÍAS CALENDARIO)")
            ws.cell(row=current_row, column=3).font = subtotal_font
            ws.cell(row=current_row, column=3).alignment = center_align
            ws.cell(row=current_row, column=3).border = thin_border
            
            ws.cell(row=current_row, column=5, value=90)
            ws.cell(row=current_row, column=5).font = subtotal_font
            ws.cell(row=current_row, column=5).alignment = center_align
            ws.cell(row=current_row, column=5).border = thin_border
            current_row += 3
            
            # Firma del representante legal
            ws.merge_cells(f'A{current_row}:C{current_row+3}')
            ws.cell(row=current_row, column=1, value="FIRMA DEL REPRESENTANTE LEGAL")
            ws.cell(row=current_row, column=1).font = subtotal_font
            ws.cell(row=current_row, column=1).alignment = left_align
            ws.cell(row=current_row, column=1).border = thin_border
            
            # Crear hoja AIU con formato detallado
            self._create_aiu_sheet(wb)
            
            # Crear hoja ANALISIS UNITARIOS con formato detallado
            self._create_analisis_unitarios_sheet(wb)
            
            # Crear hoja INSUMOS con consolidado de recursos
            self._create_insumos_sheet(wb)
            
            # Guardar archivo
            wb.save(filePath)
            
            QMessageBox.information(self, "Éxito", f"Excel exportado exitosamente como: {filePath}")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al exportar Excel: {str(e)}")
    
    def _num_a_letras_es(self, numero):
        """Convierte un número a letras en español"""
        if numero == 0:
            return "CERO PESOS"
        
        unidades = ["", "UN", "DOS", "TRES", "CUATRO", "CINCO", "SEIS", "SIETE", "OCHO", "NUEVE"]
        decenas = ["", "", "VEINTE", "TREINTA", "CUARENTA", "CINCUENTA", "SESENTA", "SETENTA", "OCHENTA", "NOVENTA"]
        centenas = ["", "CIENTO", "DOSCIENTOS", "TRESCIENTOS", "CUATROCIENTOS", "QUINIENTOS", "SEISCIENTOS", "SETECIENTOS", "OCHOCIENTOS", "NOVECIENTOS"]
        
        def convertir_centenas(n):
            if n == 0:
                return ""
            elif n == 100:
                return "CIEN"
            elif n < 10:
                return unidades[n]
            elif n < 20:
                especiales = ["DIEZ", "ONCE", "DOCE", "TRECE", "CATORCE", "QUINCE", "DIECISEIS", "DIECISIETE", "DIECIOCHO", "DIECINUEVE"]
                return especiales[n - 10]
            elif n < 30:
                return "VEINTI" + unidades[n - 20] if n > 20 else "VEINTE"
            elif n < 100:
                return decenas[n // 10] + (" Y " + unidades[n % 10] if n % 10 != 0 else "")
            else:
                return centenas[n // 100] + (" " + convertir_centenas(n % 100) if n % 100 != 0 else "")
        
        def convertir_miles(n):
            if n == 0:
                return ""
            elif n == 1:
                return "MIL"
            elif n < 1000:
                return convertir_centenas(n) + " MIL"
            else:
                return convertir_centenas(n // 1000) + " MIL " + convertir_centenas(n % 1000)
        
        def convertir_millones(n):
            if n == 0:
                return ""
            elif n == 1:
                return "UN MILLON"
            elif n < 1000:
                return convertir_centenas(n) + " MILLONES"
            else:
                return convertir_miles(n) + " MILLONES"
        
        # Convertir el número
        entero = int(numero)
        
        if entero >= 1000000:
            millones = entero // 1000000
            resto = entero % 1000000
            resultado = convertir_millones(millones)
            if resto >= 1000:
                miles = resto // 1000
                centenas_resto = resto % 1000
                if miles > 0:
                    resultado += " " + convertir_miles(miles)
                if centenas_resto > 0:
                    resultado += " " + convertir_centenas(centenas_resto)
            elif resto > 0:
                resultado += " " + convertir_centenas(resto)
        elif entero >= 1000:
            miles = entero // 1000
            resto = entero % 1000
            resultado = convertir_miles(miles)
            if resto > 0:
                resultado += " " + convertir_centenas(resto)
        else:
            resultado = convertir_centenas(entero)
        
        return resultado.strip() + " PESOS"

    def import_csv(self):
        """Importa datos desde un archivo CSV al presupuesto y calcula automáticamente los totales."""
        filePath, _ = QFileDialog.getOpenFileName(self, "Importar CSV", "", "CSV Files (*.csv)")
        if not filePath:
            return

        self.table.setRowCount(0)
        self.chapter_counter = 0

        # Bloquear señales para evitar procesamientos intermedios (on_cell_changed, etc.)
        self.table.blockSignals(True)

        total_rows_processed = 0
        analysis_rows_added = 0

        with open(filePath, 'r', encoding='utf-8') as file:
            reader = csv.reader(file)
            headers = next(reader, None)
            
            aiu_dict = {}
            analysis_rows = []
            sub_items = []
            is_aiu_block = False
            processing_sub_items = False

            for row in reader:
                total_rows_processed += 1
                if not row:
                    is_aiu_block = True
                    continue
                
                # Detectar secciones especiales
                if len(row) >= 1:
                    if row[0] == '=== SUB-ITEMS AIU ===':
                        processing_sub_items = True
                        continue
                    elif row[0] == '=== DATOS AIU ===':
                        processing_sub_items = False
                        is_aiu_block = True
                        continue
                
                # Procesar sub-items AIU
                if processing_sub_items and len(row) >= 4 and row[0].startswith('SUB_ITEM_'):
                    try:
                        sub_items.append({
                            'concepto': row[1],
                            'pct': float(row[2]) if row[2] else 0.0,
                            'valor': float(row[3]) if row[3] else 0.0
                        })
                    except (ValueError, IndexError):
                        pass
                    continue
                
                if is_aiu_block:
                    # Procesar líneas AIU nuevas y antiguas
                    if len(row) >= 2 and row[0].startswith('AIU_'):
                        key = row[0]
                        val = row[1]
                        try:
                            val_f = float(val.replace(',', '').replace('$', '').replace('%', '').strip())
                        except:
                            val_f = 0.0
                        
                        if key == "AIU_ADMIN": aiu_dict['admin'] = val_f
                        elif key == "AIU_SUB_TOTAL": aiu_dict['sub_total'] = val_f
                        elif key == "AIU_IMPREV": aiu_dict['imprev'] = val_f
                        elif key == "AIU_UTIL": aiu_dict['util'] = val_f
                        elif key == "AIU_IVA": aiu_dict['iva'] = val_f
                        elif key == "AIU_IMPREV_PCT": aiu_dict['imprev_pct'] = val_f
                        elif key == "AIU_UTIL_PCT": aiu_dict['util_pct'] = val_f
                        elif key == "AIU_IVA_PCT": aiu_dict['iva_pct'] = val_f
                        elif key == "AIU_TOTAL": aiu_dict['total_aiu'] = val_f
                    else:
                            # Formato anterior de AIU
                        key = row[0].strip().upper()
                        value = row[1].replace('$', '').replace(',', '').strip() if len(row) > 1 else '0'
                        pct = row[2].replace('%', '').strip() if len(row) > 2 else '0'
                        try:
                            val_f = float(value)
                            pct_f = float(pct)
                        except ValueError:
                            val_f = 0.0; pct_f = 0.0
                    
                    if key == "COSTO DIRECTO": aiu_dict['direct_cost'] = val_f
                    elif key == "ADMINISTRACIÓN": aiu_dict['admin'] = val_f; aiu_dict['admin_pct'] = pct_f
                    elif key == "IMPREVISTOS": aiu_dict['imprev'] = val_f; aiu_dict['imprev_pct'] = pct_f
                    elif key == "UTILIDAD": aiu_dict['util'] = val_f; aiu_dict['util_pct'] = pct_f
                    elif key == "IVA UTILIDAD": aiu_dict['iva'] = val_f; aiu_dict['iva_pct'] = pct_f
                    elif key == "TOTAL COSTOS INDIRECTOS": aiu_dict['total_aiu'] = val_f
                else:
                    analysis_rows.append(row)
            
            # Agregar sub-items al AIU dict si existen
            if sub_items:
                aiu_dict['sub_items'] = sub_items

            for row_data in analysis_rows:
                # Lógica para añadir capítulos y análisis
                if (len(row_data) >= 7 and row_data[6].strip().lower() == 'chapter') or (len(row_data) >= 6 and row_data[5].strip().lower() == 'chapter') or ('cap' in row_data[0].lower() and row_data[0][0].isdigit()):
                    chapter_text = row_data[0]
                    if '.' in chapter_text: 
                        chapter_name = chapter_text.split('.',1)[1].strip()
                    else: 
                        chapter_name = chapter_text
                    self.add_chapter_row(chapter_name, trigger_rebuild=False)
                    continue
                
                if len(row_data) < 5: 
                    continue
                
                row_idx = self.table.rowCount()
                self.table.insertRow(row_idx)
                for col in range(5):
                    val = row_data[col] if col < len(row_data) else ''
                    item = QTableWidgetItem(val)
                    if col in (2,3): 
                        item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsEditable)
                    else: 
                        item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled)
                    if col==1: 
                        item.setToolTip(val)
                    self.table.setItem(row_idx, col, item)
                
                # **CLAVE**: Restaurar el código del análisis en UserRole si existe
                codigo_analisis = row_data[5] if len(row_data) > 5 else ""
                tipo_fila = row_data[6] if len(row_data) > 6 else ""
                
                if codigo_analisis and tipo_fila == 'analysis':
                    # Es un análisis real, guardar el código en UserRole
                    item_codigo = self.table.item(row_idx, 0)
                    if item_codigo:
                        item_codigo.setData(Qt.ItemDataRole.UserRole, codigo_analisis)
                        print(f"✅ Código {codigo_analisis} restaurado en fila {row_idx}")
                
                total_item = QTableWidgetItem()
                total_item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled)
                self.table.setItem(row_idx, 5, total_item)
                self.update_row_total(row_idx)
                analysis_rows_added += 1

        if aiu_dict:
            self.aiu_breakdown = aiu_dict
            self.admin_cost_total = aiu_dict.get('total_aiu', 0.0)

        # Reconstruir tabla con subtotales
        print(f"\n=== RESUMEN IMPORTACIÓN ===")
        print(f"Total filas procesadas del CSV: {total_rows_processed}")
        print(f"Análisis agregados a la tabla: {analysis_rows_added}")
        print(f"Filas totales en tabla antes de rebuild: {self.table.rowCount()}")
        
        # Verificar costos unitarios antes del rebuild
        print("\n=== VERIFICACIÓN COSTOS ANTES DE REBUILD ===")
        self._dump_table("DUMP ANTES REBUILD")
        
        self.rebuild_table_safe()
        
        # Desbloquear señales ahora que la importación ha terminado
        self.table.blockSignals(False)
        
        # Verificar costos unitarios después de rebuild
        print("\n=== VERIFICACIÓN COSTOS DESPUÉS DE REBUILD ===")
        self._dump_table("DUMP DESPUÉS REBUILD")
        
        print(f"Filas totales en tabla después de rebuild: {self.table.rowCount()}")
        QMessageBox.information(self, "Importado", f"El presupuesto ha sido importado desde CSV.\nAnálisis importados: {analysis_rows_added}\nTotales calculados automáticamente.")

    def renumber_items(self):
        """Renumera todos los capítulos y los ítems de la tabla, omitiendo subtotales."""
        # Primero, renumerar capítulos
        chapter_count = 0
        for row in range(self.table.rowCount()):
            item = self.table.item(row, 0)
            if item and item.data(Qt.ItemDataRole.UserRole) == 'chapter':
                chapter_count += 1
                try:
                    # Extraer solo el nombre, sin el número anterior
                    name = item.text().split('.', 1)[1].strip()
                except IndexError:
                    name = item.text()
                item.setText(f"{chapter_count}. {name}")
                item.setData(Qt.ItemDataRole.UserRole + 1, chapter_count)
        self.chapter_counter = chapter_count

        # Segundo, renumerar ítems de análisis (omitiendo subtotales)
        current_chapter_number = 0
        item_in_chapter_counter = 0
        for row in range(self.table.rowCount()):
            item = self.table.item(row, 0)
            if not item:
                continue
                
            if item.data(Qt.ItemDataRole.UserRole) == 'chapter':
                current_chapter_number = item.data(Qt.ItemDataRole.UserRole + 1)
                item_in_chapter_counter = 0  # Reset for new chapter
            elif item.data(Qt.ItemDataRole.UserRole) == 'subtotal':
                # Omitir filas de subtotal
                continue
            else:
                # Es un ítem de análisis
                item_in_chapter_counter += 1
                item.setText(f"{current_chapter_number}.{item_in_chapter_counter}")

    def rebuild_table(self):
        """Reconstruye completamente la tabla eliminando subtotales y recalculando todo."""
        # Bloquear señales durante la reconstrucción
        self.table.blockSignals(True)
        
        try:
            # Renumerar todos los items primero
            self.renumber_items()
            
            # NO recalcular los totales de fila durante la importación
            # Solo calcular los totales si los costos unitarios ya están establecidos
            # Esto evita sobrescribir datos recién importados
            
            # Finalmente, actualizar totales y subtotales
            self.update_total_presupuesto()
                    
        finally:
            self.table.blockSignals(False)

    def rebuild_table_safe(self):
        """Versión segura de rebuild que no modifica costos unitarios existentes."""
        # Bloquear señales durante la reconstrucción
        self.table.blockSignals(True)
        
        try:
            # Renumerar todos los items primero
            self.renumber_items()
            
            # Solo actualizar totales y subtotales, sin tocar costos unitarios
            self.update_total_presupuesto()
            
        finally:
            self.table.blockSignals(False)

    # --- MÉTODO AUXILIAR DE DEPURACIÓN ---
    def _dump_table(self, label="DUMP"):
        """Imprime en consola el contenido actual de la tabla (solo columnas clave)."""
        print(f"\n--- {label} ---")
        for row in range(self.table.rowCount()):
            item = self.table.item(row, 0)
            desc = self.table.item(row, 1)
            cu = self.table.item(row, 4)
            ct = self.table.item(row, 5)
            if item:
                item_text = item.text()
                cu_text = cu.text() if cu else ""
                ct_text = ct.text() if ct else ""
                print(f"Fila {row:2d} | Item={item_text!r} | CU={cu_text!r} | CT={ct_text!r}")

    # ---------------- NUEVO MÉTODO: Editar análisis seleccionado ------------------
    def edit_selected_analysis(self):
        """Emite la señal de edición de análisis para la fila seleccionada."""
        selected_items = self.table.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Sin selección", "Seleccione una fila de análisis para editar.")
            return

        row = selected_items[0].row()
        item = self.table.item(row, 0)
        if not item:
            QMessageBox.warning(self, "Error", "La fila seleccionada no es válida.")
            return

        user_role = item.data(Qt.ItemDataRole.UserRole)
        if user_role in (None, "", "chapter", "subtotal"):
            QMessageBox.warning(self, "Operación no válida", "Solo se pueden editar filas de análisis (no capítulos ni subtotales).")
            return

        self.analysis_edit_requested.emit(user_role)

    def open_administracion_window(self):
        """Abre la ventana para calcular costos administrativos basados en profesionales."""
        # Obtener los profesionales de la base de datos
        session = SessionLocal()
        try:
            profesionales_db = session.query(Profesional).all()
            profesionales = [
                {
                    'nombre': p.nombre,
                    'cargo': p.cargo,
                    'salario_mensual': p.salario_mensual,
                    'necesario': p.necesario,
                } for p in profesionales_db
            ]
        finally:
            session.close()

        # Mostrar la vista
        direct_cost = getattr(self, 'direct_cost_total', 0.0)
        self.admin_dialog = AdministracionWindow(profesionales, direct_cost, parent=self)
        self.admin_dialog.aiu_computed.connect(self.on_admin_cost_computed)
        self.admin_dialog.exec()

    def on_admin_cost_computed(self, data):
        """Recibe desglose AIU y actualiza totales."""
        # data dict includes totals
        self.admin_cost_total = data.get('total_aiu', 0.0)
        self.aiu_breakdown = data
        self.update_total_presupuesto()
    
    def _create_aiu_sheet(self, workbook):
        """Crea la hoja de AIU con formato detallado"""
        ws = workbook.create_sheet("AIU")
        
        # Estilos
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        header_font = Font(bold=True, size=12)
        normal_font = Font(size=10)
        small_font = Font(size=9)
        
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Header
        ws.merge_cells('A1:C1')
        ws['A1'] = "NOMBRE DE LA EMPRESA"
        ws['A1'].font = header_font
        ws['A1'].alignment = center_align
        ws['A1'].border = thin_border
        
        ws.merge_cells('D1:G1')
        ws['D1'] = "LOGO DE LA EMPRESA"
        ws['D1'].font = header_font
        ws['D1'].alignment = center_align
        ws['D1'].border = thin_border
        
        ws.row_dimensions[1].height = 25
        
        # Título
        ws.merge_cells('A3:G3')
        ws['A3'] = "DISCRIMINACIÓN DEL AIU"
        ws['A3'].font = Font(bold=True, size=14)
        ws['A3'].alignment = center_align
        ws['A3'].fill = header_fill
        ws['A3'].border = thin_border
        
        # Información básica
        ws['A4'] = "Obra:"
        ws.merge_cells('B4:E4')
        ws['B4'] = "ESCRIBA AQUÍ EL NOMBRE DE LA OBRA"
        ws['B4'].border = thin_border
        
        ws['F4'] = "FECHA:"
        ws['G4'] = "19-sept-25"
        ws['G4'].border = thin_border
        
        ws['F5'] = "QUIEN ELABORÓ:"
        ws['G5'].border = thin_border
        
        # Obtener datos del AIU
        direct_cost = sum(ch['subtotal'] for ch in self._get_chapters_data())
        
        # Intentar obtener datos del AIU embebido en MainWindow
        aiu_data = None
        aiu_source = "ninguna"
        
        try:
            # Buscar la ventana principal
            from PyQt6.QtWidgets import QApplication
            app = QApplication.instance()
            if app:
                for widget in app.topLevelWidgets():
                    if hasattr(widget, 'aiu_widget') and widget.aiu_widget:
                        aiu_data = self._extract_aiu_data_from_widget(widget.aiu_widget)
                        aiu_source = "aiu_widget embebido"
                        break
        except Exception as e:
            pass
        
        # Si no se encontró AIU embebido, usar admin_dialog
        if not aiu_data:
            admin_window = getattr(self, 'admin_dialog', None)
            if admin_window:
                aiu_data = self._extract_aiu_data_from_widget(admin_window)
                aiu_source = "admin_dialog"
        
        # Si aún no hay datos, usar valores por defecto
        if not aiu_data:
            aiu_data = {
                'costo_directo': direct_cost,
                'admin_total': direct_cost * 0.34,
                'imprev_total': direct_cost * 0.05,
                'util_total': direct_cost * 0.05,
                'iva_total': direct_cost * 0.0058,
                'profesionales': [],
                'oficina': [],
                'polizas': [],
                'estampillas': []
            }
        
        current_row = 7
        
        # Tabla principal de costos
        ws.merge_cells('A7:G7')
        ws.cell(row=7, column=1, value="COSTOS ADMINISTRATIVOS")
        ws.cell(row=7, column=1).font = header_font
        ws.cell(row=7, column=1).fill = green_fill
        ws.cell(row=7, column=1).alignment = center_align
        
        # Headers
        headers = ["", "VALOR MENSUAL", "%DEDIC", "MESES", "VALOR PARCIAL", "% SOBRE CD", "VALOR EN PESOS"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=8, column=col, value=header)
            ws.cell(row=8, column=col).font = small_font
            ws.cell(row=8, column=col).fill = light_fill
            ws.cell(row=8, column=col).alignment = center_align
            ws.cell(row=8, column=col).border = thin_border
        
        current_row = 9
        
        # Personal técnico
        ws.cell(row=current_row, column=1, value="PERSONAL TÉCNICO Y ADM. (Incl. Prest. Soc.)")
        ws.cell(row=current_row, column=1).font = small_font
        ws.cell(row=current_row, column=1).alignment = left_align
        ws.cell(row=current_row, column=1).border = thin_border
        current_row += 1
        
        # Agregar profesionales
        total_profesionales = 0
        profesionales_list = aiu_data.get('profesionales', [])
        
        # Si no hay profesionales, mostrar mensaje informativo
        if not profesionales_list:
            ws.cell(row=current_row, column=1, value="(No se encontraron profesionales configurados)")
            ws.cell(row=current_row, column=1).font = small_font
            ws.cell(row=current_row, column=1).alignment = left_align
            for col in range(1, 8):
                ws.cell(row=current_row, column=col).border = thin_border
            current_row += 1
        else:
            # Agregar profesionales reales
            for prof in profesionales_list:
                profesional = prof.get('profesional', '')
                cargo = prof.get('cargo', '')
                salario = prof.get('salario_mensual', 0)
                dedicacion = prof.get('dedicacion', 0)
                meses = prof.get('meses', 6)
                valor_total = prof.get('valor_total', 0)
                
                # Mostrar solo el cargo (que contiene la información principal)
                nombre_a_mostrar = cargo if cargo else profesional
                ws.cell(row=current_row, column=1, value=nombre_a_mostrar)
                ws.cell(row=current_row, column=2, value=salario)
                ws.cell(row=current_row, column=3, value=f"{dedicacion:.1f}%")
                ws.cell(row=current_row, column=4, value=meses)
                
                valor_parcial = salario * meses * (dedicacion/100) if salario > 0 and meses > 0 and dedicacion > 0 else valor_total
                ws.cell(row=current_row, column=5, value=valor_parcial)
                
                porcentaje_cd = (valor_parcial / aiu_data['costo_directo'] * 100) if aiu_data['costo_directo'] > 0 else 0
                ws.cell(row=current_row, column=6, value=f"{porcentaje_cd:.2f}%")
                ws.cell(row=current_row, column=7, value=valor_parcial)
                
                total_profesionales += valor_parcial
                
                # Aplicar formato
                for col in range(1, 8):
                    ws.cell(row=current_row, column=col).font = small_font
                    ws.cell(row=current_row, column=col).border = thin_border
                    if col in [2, 5, 7]:
                        ws.cell(row=current_row, column=col).number_format = '#,##0.00'
                        ws.cell(row=current_row, column=col).alignment = right_align
                    else:
                        ws.cell(row=current_row, column=col).alignment = center_align
                
                current_row += 1
        
        # Subtotal profesionales
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws.cell(row=current_row, column=1, value="TOTAL COSTOS ADMINISTRATIVOS (SUMA):")
        ws.cell(row=current_row, column=1).font = header_font
        ws.cell(row=current_row, column=1).alignment = right_align
        ws.cell(row=current_row, column=7, value=total_profesionales)
        ws.cell(row=current_row, column=7).font = header_font
        ws.cell(row=current_row, column=7).number_format = '#,##0.00'
        ws.cell(row=current_row, column=7).alignment = right_align
        
        for col in range(1, 8):
            ws.cell(row=current_row, column=col).border = thin_border
            ws.cell(row=current_row, column=col).fill = light_fill
        
        current_row += 3
        
        # OFICINA, PAPELERÍA Y OTROS
        ws.merge_cells(f'A{current_row}:G{current_row}')
        ws.cell(row=current_row, column=1, value="OFICINA, PAPELERÍA Y OTROS:")
        ws.cell(row=current_row, column=1).font = header_font
        ws.cell(row=current_row, column=1).fill = green_fill
        ws.cell(row=current_row, column=1).alignment = center_align
        ws.cell(row=current_row, column=1).border = thin_border
        current_row += 1
        
        # Headers oficina
        for col, header in enumerate(headers, 1):
            ws.cell(row=current_row, column=col, value=header)
            ws.cell(row=current_row, column=col).font = small_font
            ws.cell(row=current_row, column=col).fill = light_fill
            ws.cell(row=current_row, column=col).alignment = center_align
            ws.cell(row=current_row, column=col).border = thin_border
        current_row += 1
        
        # Datos de oficina (extraer de AIU real)
        oficina_items = aiu_data.get('oficina', [])
        if not oficina_items:
            oficina_items = [
                {"concepto": "COSTO PAPELERÍA, FOTOCOPIAS, ETC.", "valor_base": 400000, "dedicacion": 100, "meses": 6, "valor": 464200},
                {"concepto": "COSTO OFICINA", "valor_base": 500000, "dedicacion": 70, "meses": 6, "valor": 364200}
            ]
        
        total_oficina = 0
        for item_data in oficina_items:
            if isinstance(item_data, dict):
                item = item_data.get('concepto', '')
                valor_mensual = item_data.get('valor_base', 0)
                dedic = item_data.get('dedicacion', 0)
                meses = item_data.get('meses', 6)
                valor_final = item_data.get('valor', 0)
            else:
                # Formato antiguo (tupla)
                item, valor_mensual, dedic, meses, pct_cd, valor_final = item_data
            
            ws.cell(row=current_row, column=1, value=item)
            ws.cell(row=current_row, column=2, value=valor_mensual)
            ws.cell(row=current_row, column=3, value=f"{dedic}%")
            ws.cell(row=current_row, column=4, value=meses)
            
            valor_parcial = valor_mensual * meses * (dedic/100)
            ws.cell(row=current_row, column=5, value=valor_parcial)
            
            # Calcular porcentaje sobre costo directo
            pct_cd = (valor_final / aiu_data['costo_directo'] * 100) if aiu_data['costo_directo'] > 0 else 0
            ws.cell(row=current_row, column=6, value=f"{pct_cd:.2f}%")
            ws.cell(row=current_row, column=7, value=valor_final)
            
            total_oficina += valor_final
            
            for col in range(1, 8):
                ws.cell(row=current_row, column=col).font = small_font
                ws.cell(row=current_row, column=col).border = thin_border
                if col in [2, 5, 7]:
                    ws.cell(row=current_row, column=col).number_format = '#,##0.00'
                    ws.cell(row=current_row, column=col).alignment = right_align
                else:
                    ws.cell(row=current_row, column=col).alignment = center_align
            
            current_row += 1
        
        current_row += 2
        
        # LEGALIZACIÓN DEL CONTRATO (Pólizas)
        ws.merge_cells(f'A{current_row}:G{current_row}')
        ws.cell(row=current_row, column=1, value="LEGALIZACIÓN DEL CONTRATO (Pólizas):")
        ws.cell(row=current_row, column=1).font = header_font
        ws.cell(row=current_row, column=1).fill = green_fill
        ws.cell(row=current_row, column=1).alignment = center_align
        ws.cell(row=current_row, column=1).border = thin_border
        current_row += 1
        
        # Headers pólizas
        polizas_headers = ["", "%Requerido", "Duración", "%Prima", "Vr. TI Contrato", "%Dedic Directos", "Valor en Pesos"]
        for col, header in enumerate(polizas_headers, 1):
            ws.cell(row=current_row, column=col, value=header)
            ws.cell(row=current_row, column=col).font = small_font
            ws.cell(row=current_row, column=col).fill = light_fill
            ws.cell(row=current_row, column=col).alignment = center_align
            ws.cell(row=current_row, column=col).border = thin_border
        current_row += 1
        
        # Datos de pólizas (extraer de AIU real)
        polizas_items = aiu_data.get('polizas', [])
        if not polizas_items:
            polizas_items = [
                {"concepto": "PÓLIZA DE CUMPLIMIENTO", "requerido": 20, "duracion": 10, "prima": 0.50, "vr_ti": 2795000, "dedicacion": 4.87, "valor": 71654400},
                {"concepto": "PÓLIZA DE CALIDAD DEL ANTICIPO", "requerido": 100, "duracion": 10, "prima": 0.50, "vr_ti": 5590000, "dedicacion": 7.31, "valor": 5590000}
            ]
        
        total_polizas = 0
        for item_data in polizas_items:
            if isinstance(item_data, dict):
                item = item_data.get('concepto', '')
                req = item_data.get('requerido', 0)
                duracion = item_data.get('duracion', 0)
                prima = item_data.get('prima', 0)
                vr_ti = item_data.get('vr_ti', 0)
                dedic = item_data.get('dedicacion', 0)
                valor_final = item_data.get('valor', 0)
            else:
                # Formato antiguo (tupla)
                item, req, duracion, prima, vr_ti, dedic, valor_final = item_data
            ws.cell(row=current_row, column=1, value=item)
            ws.cell(row=current_row, column=2, value=f"{req:.0f}%")
            ws.cell(row=current_row, column=3, value=duracion)
            ws.cell(row=current_row, column=4, value=f"{prima:.2f}%")
            ws.cell(row=current_row, column=5, value=vr_ti)
            ws.cell(row=current_row, column=6, value=f"{dedic:.2f}%")
            ws.cell(row=current_row, column=7, value=valor_final)
            
            total_polizas += valor_final
            
            for col in range(1, 8):
                ws.cell(row=current_row, column=col).font = small_font
                ws.cell(row=current_row, column=col).border = thin_border
                if col in [5, 7]:
                    ws.cell(row=current_row, column=col).number_format = '#,##0.00'
                    ws.cell(row=current_row, column=col).alignment = right_align
                else:
                    ws.cell(row=current_row, column=col).alignment = center_align
            
            current_row += 1
        
        current_row += 2
        
        # DESCUENTOS SOBRE EL CONTRATO (Estampillas)
        ws.merge_cells(f'A{current_row}:D{current_row}')
        ws.cell(row=current_row, column=1, value="DESCUENTOS SOBRE EL CONTRATO:")
        ws.cell(row=current_row, column=1).font = header_font
        ws.cell(row=current_row, column=1).fill = green_fill
        ws.cell(row=current_row, column=1).alignment = center_align
        ws.cell(row=current_row, column=1).border = thin_border
        current_row += 1
        
        # Headers estampillas (solo 3 columnas)
        estampillas_headers = ["", "Vr. TI Contrato", "%", "Valor en Pesos"]
        for col, header in enumerate(estampillas_headers, 1):
            if col <= 4:  # Solo las primeras 4 columnas (A, B, C, D)
                ws.cell(row=current_row, column=col, value=header)
                ws.cell(row=current_row, column=col).font = small_font
                ws.cell(row=current_row, column=col).fill = light_fill
                ws.cell(row=current_row, column=col).alignment = center_align
                ws.cell(row=current_row, column=col).border = thin_border
        current_row += 1
        
        # Datos de estampillas (extraer de AIU real)
        estampillas_items = aiu_data.get('estampillas', [])
        if not estampillas_items:
            estampillas_items = [
                {"concepto": "Estampilla pro Desarrollo", "vr_ti": 374052564, "porcentaje": 4, "valor": 72742543},
                {"concepto": "Estampilla pro Univalle", "vr_ti": 2078069804, "porcentaje": 1.43, "valor": 20780698},
                {"concepto": "Estampilla pro Hospital", "vr_ti": 2078069804, "porcentaje": 1, "valor": 20780698}
            ]
        
        total_estampillas = 0
        for item_data in estampillas_items:
            if isinstance(item_data, dict):
                item = item_data.get('concepto', '')
                vr_ti = item_data.get('vr_ti', 0)
                pct = item_data.get('porcentaje', 0)
                valor_final = item_data.get('valor', 0)
                col2 = col3 = col5 = ""
            else:
                # Formato antiguo (tupla)
                item, col2, col3, vr_ti, col5, pct, valor_final = item_data
            # Solo usar 4 columnas: Concepto, Vr. TI Contrato, %, Valor en Pesos
            ws.cell(row=current_row, column=1, value=item)
            ws.cell(row=current_row, column=2, value=vr_ti)
            ws.cell(row=current_row, column=3, value=f"{pct}%")
            ws.cell(row=current_row, column=4, value=valor_final)
            
            total_estampillas += valor_final
            
            for col in range(1, 5):  # Solo columnas A, B, C, D
                ws.cell(row=current_row, column=col).font = small_font
                ws.cell(row=current_row, column=col).border = thin_border
                if col in [2, 4]:  # Vr. TI Contrato y Valor en Pesos
                    ws.cell(row=current_row, column=col).number_format = '#,##0.00'
                    ws.cell(row=current_row, column=col).alignment = right_align
                elif col == 3:  # Porcentaje
                    ws.cell(row=current_row, column=col).alignment = center_align
                else:  # Concepto
                    ws.cell(row=current_row, column=col).alignment = left_align
            
            current_row += 1
        
        current_row += 2
        
        # TOTALES FINALES - Usar datos reales del AIU
        costo_directo = aiu_data['costo_directo']
        admin_total = aiu_data.get('admin_total', 0)
        imprev_total = aiu_data.get('imprev_total', 0)
        util_total = aiu_data.get('util_total', 0)
        iva_total = aiu_data.get('iva_total', 0)
        
        # Calcular porcentajes sobre costo directo
        admin_pct = (admin_total / costo_directo * 100) if costo_directo > 0 else 0
        imprev_pct = (imprev_total / costo_directo * 100) if costo_directo > 0 else 0
        util_pct = (util_total / costo_directo * 100) if costo_directo > 0 else 0
        
        # Para IVA, calcular sobre la utilidad
        iva_base = util_total
        iva_pct = (iva_total / iva_base * 100) if iva_base > 0 else 0
        
        total_aiu = admin_total + imprev_total + util_total + iva_total
        total_aiu_pct = (total_aiu / costo_directo * 100) if costo_directo > 0 else 0
        
        # Tabla de resumen final
        resumen_totales = [
            ("ADMINISTRACIÓN:", f"{admin_pct:.2f}%", admin_total, True),
            ("IMPREVISTOS:", f"{imprev_pct:.2f}%", imprev_total, False),
            ("UTILIDAD:", f"{util_pct:.2f}%", util_total, False),
            ("IVA SOBRE LA UTILIDAD:", f"{iva_pct:.2f}%", iva_total, False),
            ("COSTO TOTAL DEL AIU:", f"{total_aiu_pct:.2f}%", total_aiu, True)
        ]
        
        for label, pct, valor, is_bold in resumen_totales:
            ws.merge_cells(f'A{current_row}:C{current_row}')
            ws.cell(row=current_row, column=1, value=label)
            ws.cell(row=current_row, column=1).font = header_font if is_bold else normal_font
            ws.cell(row=current_row, column=1).alignment = right_align
            
            ws.cell(row=current_row, column=4, value=pct)
            ws.cell(row=current_row, column=4).font = header_font if is_bold else normal_font
            ws.cell(row=current_row, column=4).alignment = center_align
            
            ws.cell(row=current_row, column=5, value=valor)
            ws.cell(row=current_row, column=5).font = header_font if is_bold else normal_font
            ws.cell(row=current_row, column=5).number_format = '#,##0.00'
            ws.cell(row=current_row, column=5).alignment = right_align
            
            # Aplicar bordes y fondo para filas importantes
            for col in range(1, 6):
                ws.cell(row=current_row, column=col).border = thin_border
                if is_bold:
                    ws.cell(row=current_row, column=col).fill = light_fill
            
            current_row += 1
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 15
    
    def _get_chapters_data(self):
        """Obtiene datos de capítulos para cálculos"""
        from PyQt6.QtCore import Qt
        chapters = []
        current = None
        
        for tr in range(self.table.rowCount()):
            it0 = self.table.item(tr, 0)
            if not it0:
                continue
                
            role = it0.data(Qt.ItemDataRole.UserRole)
            
            if role == 'chapter':
                if current:
                    chapters.append(current)
                desc1 = self.table.item(tr, 1)
                current = {
                    'title': it0.text(),
                    'desc': desc1.text() if desc1 else '',
                    'items': [],
                    'subtotal': 0.0
                }
            elif role == 'subtotal':
                val_item = self.table.item(tr, 5)
                val_txt = val_item.text() if val_item else '0'
                try:
                    st = float(str(val_txt).replace('$','').replace(',','').replace(' ',''))
                except Exception:
                    st = 0.0
                if current:
                    current['subtotal'] = st
        
        if current:
            chapters.append(current)
        return chapters
    
    def _extract_aiu_data_from_widget(self, admin_window):
        """Extrae datos detallados del AIU window"""
        data = {
            'costo_directo': admin_window.costo_directo,
            'admin_total': getattr(admin_window, '_admin_total', 0),
            'imprev_total': getattr(admin_window, '_imprev_total', 0),
            'util_total': getattr(admin_window, '_util_total', 0),
            'iva_total': getattr(admin_window, '_iva_total', 0),
            'profesionales': [],
            'oficina': [],
            'polizas': [],
            'estampillas': []
        }
        
        # Extraer datos de profesionales
        if hasattr(admin_window, 'table'):
            table = admin_window.table
            for row in range(table.rowCount()):  # Incluir todas las filas
                try:
                    # Columnas: "Profesional", "Cargo", "Salario Mensual", "% Dedicación", "Meses", "Total"
                    profesional_item = table.item(row, 0)
                    cargo_item = table.item(row, 1)
                    salario_item = table.item(row, 2)
                    dedicacion_item = table.item(row, 3)
                    meses_item = table.item(row, 4)
                    valor_item = table.item(row, 5)
                    
                    profesional = profesional_item.text() if profesional_item else ""
                    cargo = cargo_item.text() if cargo_item else ""
                    
                    if profesional and not profesional.upper().startswith("SUBTOTAL") and profesional.strip() != "":
                        salario = float((salario_item.text() if salario_item else '0').replace('$','').replace(',','').strip())
                        dedicacion = float((dedicacion_item.text() if dedicacion_item else '0').replace('%','').strip())
                        meses = float((meses_item.text() if meses_item else '0').strip())
                        valor_total = float((valor_item.text() if valor_item else '0').replace('$','').replace(',','').strip())
                        
                        data['profesionales'].append({
                            'profesional': profesional,
                            'cargo': cargo,
                            'salario_mensual': salario,
                            'dedicacion': dedicacion,
                            'meses': meses,
                            'valor_total': valor_total
                        })
                except Exception as e:
                    continue
        
        # Extraer datos de oficina
        if hasattr(admin_window, 'tbl_oficina'):
            table = admin_window.tbl_oficina
            for row in range(table.rowCount()):  # Incluir todas las filas
                try:
                    concepto_item = table.item(row, 0)
                    concepto = concepto_item.text() if concepto_item else ""
                    if concepto and not concepto.upper().startswith("SUBTOTAL") and concepto.strip() != "":
                        valor_base_item = table.item(row, 1)
                        dedicacion_item = table.item(row, 2)
                        meses_item = table.item(row, 3)
                        valor_item = table.item(row, 4)
                        
                        valor_base = float((valor_base_item.text() if valor_base_item else '0').replace('$','').replace(',','').strip())
                        dedicacion = float((dedicacion_item.text() if dedicacion_item else '0').replace('%','').strip())
                        meses = float((meses_item.text() if meses_item else '0').strip())
                        valor = float((valor_item.text() if valor_item else '0').replace('$','').replace(',','').strip())
                        
                        data['oficina'].append({
                            'concepto': concepto,
                            'valor_base': valor_base,
                            'dedicacion': dedicacion,
                            'meses': meses,
                            'valor': valor
                        })
                except Exception:
                    continue
        
        # Extraer datos de pólizas
        if hasattr(admin_window, 'tbl_polizas'):
            table = admin_window.tbl_polizas
            for row in range(table.rowCount()):  # Incluir todas las filas
                try:
                    concepto_item = table.item(row, 0)
                    concepto = concepto_item.text() if concepto_item else ""
                    if concepto and not concepto.upper().startswith("SUBTOTAL") and concepto.strip() != "":
                        requerido_item = table.item(row, 1)
                        duracion_item = table.item(row, 2)
                        prima_item = table.item(row, 3)
                        vr_ti_item = table.item(row, 4)
                        dedicacion_item = table.item(row, 5)
                        valor_item = table.item(row, 6)
                        
                        requerido = float((requerido_item.text() if requerido_item else '0').replace('%','').strip())
                        duracion = float((duracion_item.text() if duracion_item else '0').strip())
                        prima = float((prima_item.text() if prima_item else '0').replace('%','').strip())
                        vr_ti = float((vr_ti_item.text() if vr_ti_item else '0').replace('$','').replace(',','').strip())
                        dedicacion = float((dedicacion_item.text() if dedicacion_item else '0').replace('%','').strip())
                        valor = float((valor_item.text() if valor_item else '0').replace('$','').replace(',','').strip())
                        
                        data['polizas'].append({
                            'concepto': concepto,
                            'requerido': requerido,
                            'duracion': duracion,
                            'prima': prima,
                            'vr_ti': vr_ti,
                            'dedicacion': dedicacion,
                            'valor': valor
                        })
                except Exception:
                    continue
        
        # Extraer datos de estampillas
        if hasattr(admin_window, 'tbl_estamp'):
            table = admin_window.tbl_estamp
            for row in range(table.rowCount()):  # Incluir todas las filas
                try:
                    concepto_item = table.item(row, 0)
                    concepto = concepto_item.text() if concepto_item else ""
                    if concepto and not concepto.upper().startswith("SUBTOTAL") and concepto.strip() != "":
                        porcentaje_item = table.item(row, 1)
                        valor_item = table.item(row, 2)
                        
                        porcentaje = float((porcentaje_item.text() if porcentaje_item else '0').replace('%','').strip())
                        valor = float((valor_item.text() if valor_item else '0').replace('$','').replace(',','').strip())
                        
                        data['estampillas'].append({
                            'concepto': concepto,
                            'vr_ti': admin_window.costo_directo,  # Usar costo directo como base
                            'porcentaje': porcentaje,
                            'valor': valor
                        })
                except Exception:
                    continue
        
        return data
    
    def _create_analisis_unitarios_sheet(self, workbook):
        """Crea la hoja de ANALISIS UNITARIOS con una tabla por cada análisis"""
        ws = workbook.create_sheet("ANALISIS UNITARIOS")
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 15  # CÓDIGO RECURSO
        ws.column_dimensions['B'].width = 40  # DESCRIPCIÓN
        ws.column_dimensions['C'].width = 8   # UND
        ws.column_dimensions['D'].width = 10  # CANT.
        ws.column_dimensions['E'].width = 10  # DESP.%
        ws.column_dimensions['F'].width = 15  # PRECIO UNIT
        ws.column_dimensions['G'].width = 15  # VALOR PARCIAL (mismo tamaño que PRECIO UNIT)
        
        # Estilos
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        header_font = Font(bold=True, size=12)
        normal_font = Font(size=10)
        small_font = Font(size=9)
        
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        blue_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Header
        ws.merge_cells('A1:C1')
        ws['A1'] = "NOMBRE DE LA EMPRESA"
        ws['A1'].font = header_font
        ws['A1'].alignment = center_align
        ws['A1'].border = thin_border
        
        ws.merge_cells('D1:F1')
        ws['D1'] = "LOGO DE LA EMPRESA"
        ws['D1'].font = header_font
        ws['D1'].alignment = center_align
        ws['D1'].border = thin_border
        
        ws.row_dimensions[1].height = 25
        
        # Título
        ws.merge_cells('A3:F3')
        ws['A3'] = "ANÁLISIS DE PRECIOS UNITARIOS"
        ws['A3'].font = Font(bold=True, size=14)
        ws['A3'].alignment = center_align
        ws['A3'].fill = header_fill
        ws['A3'].border = thin_border
        
        # Información básica
        ws['A4'] = "Obra:"
        ws.merge_cells('B4:D4')
        ws['B4'] = "ESCRIBA AQUÍ EL NOMBRE DE LA OBRA"
        ws['B4'].border = thin_border
        
        ws['E4'] = "FECHA:"
        ws['F4'] = "19-sept-25"
        ws['F4'].border = thin_border
        
        ws['E5'] = "QUIEN ELABORÓ:"
        ws['F5'].border = thin_border
        
        current_row = 7
        
        # Obtener análisis únicos del presupuesto
        analisis_list = self._get_unique_analisis_from_presupuesto()
        
        # Obtener porcentajes AIU para aplicar
        aiu_percentages = self._get_aiu_percentages()
        
        # Crear una tabla por cada análisis
        for analisis_data in analisis_list:
            current_row = self._create_single_analisis_table(ws, analisis_data, aiu_percentages, current_row, 
                                                           header_font, normal_font, small_font, 
                                                           header_fill, light_fill, blue_fill,
                                                           center_align, left_align, right_align, thin_border)
            current_row += 3  # Espacio entre tablas
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
    
    def _get_unique_analisis_from_presupuesto(self):
        """Obtiene la lista de análisis únicos del presupuesto"""
        from PyQt6.QtCore import Qt
        analisis_set = set()
        analisis_list = []
        
        # Recorrer la tabla del presupuesto para encontrar análisis únicos
        for row in range(self.table.rowCount()):
            item = self.table.item(row, 0)
            if not item:
                continue
            
            role = item.data(Qt.ItemDataRole.UserRole)
            if role and role != 'chapter' and role != 'subtotal':
                # Es un análisis
                if role not in analisis_set:
                    analisis_set.add(role)
                    desc_item = self.table.item(row, 1)
                    und_item = self.table.item(row, 2)
                    cant_item = self.table.item(row, 3)
                    
                    descripcion = desc_item.text() if desc_item else ""
                    unidad = und_item.text() if und_item else ""
                    cantidad = float((cant_item.text() or '0').replace(',','')) if cant_item else 0
                    
                    analisis_list.append({
                        'codigo': role,
                        'descripcion': descripcion,
                        'unidad': unidad,
                        'cantidad': cantidad
                    })
        
        return analisis_list
    
    def _get_aiu_percentages(self):
        """Obtiene los porcentajes AIU calculados"""
        # Obtener datos del AIU
        direct_cost = sum(ch['subtotal'] for ch in self._get_chapters_data())
        
        # Intentar obtener datos del AIU embebido
        aiu_data = None
        try:
            from PyQt6.QtWidgets import QApplication
            app = QApplication.instance()
            if app:
                for widget in app.topLevelWidgets():
                    if hasattr(widget, 'aiu_widget') and widget.aiu_widget:
                        aiu_data = self._extract_aiu_data_from_widget(widget.aiu_widget)
                        break
        except Exception:
            pass
        
        if not aiu_data:
            admin_window = getattr(self, 'admin_dialog', None)
            if admin_window:
                aiu_data = self._extract_aiu_data_from_widget(admin_window)
        
        if not aiu_data:
            # Valores por defecto
            return {
                'admin_pct': 5.08,
                'imprev_pct': 5.00,
                'util_pct': 5.00,
                'iva_pct': 19.00,
                'total_indirectos_pct': 6.18
            }
        
        # Calcular porcentajes reales
        costo_directo = aiu_data['costo_directo']
        admin_total = aiu_data.get('admin_total', 0)
        imprev_total = aiu_data.get('imprev_total', 0)
        util_total = aiu_data.get('util_total', 0)
        iva_total = aiu_data.get('iva_total', 0)
        
        admin_pct = (admin_total / costo_directo * 100) if costo_directo > 0 else 0
        imprev_pct = (imprev_total / costo_directo * 100) if costo_directo > 0 else 0
        util_pct = (util_total / costo_directo * 100) if costo_directo > 0 else 0
        iva_pct = (iva_total / util_total * 100) if util_total > 0 else 0
        
        total_aiu = admin_total + imprev_total + util_total + iva_total
        total_indirectos_pct = (total_aiu / costo_directo * 100) if costo_directo > 0 else 0
        
        return {
            'admin_pct': admin_pct,
            'imprev_pct': imprev_pct,
            'util_pct': util_pct,
            'iva_pct': iva_pct,
            'total_indirectos_pct': total_indirectos_pct
        }
    
    def _create_single_analisis_table(self, ws, analisis_data, aiu_percentages, start_row,
                                    header_font, normal_font, small_font,
                                    header_fill, light_fill, blue_fill,
                                    center_align, left_align, right_align, thin_border):
        """Crea una tabla individual para un análisis unitario"""
        current_row = start_row
        
        # Header del análisis
        ws.merge_cells(f'A{current_row}:G{current_row}')
        ws.cell(row=current_row, column=1, value=f"{analisis_data['codigo']} - {analisis_data['descripcion']}")
        ws.cell(row=current_row, column=1).font = header_font
        ws.cell(row=current_row, column=1).alignment = left_align
        ws.cell(row=current_row, column=1).fill = blue_fill
        for col in range(1, 8):  # Aplicar bordes a todas las columnas (A-G)
            ws.cell(row=current_row, column=col).border = thin_border
        current_row += 1
        
        # Información del análisis
        ws.cell(row=current_row, column=1, value="UNIDAD:")
        ws.cell(row=current_row, column=2, value=analisis_data['unidad'])
        ws.cell(row=current_row, column=4, value="ITEM:")
        ws.cell(row=current_row, column=5, value=analisis_data['codigo'])
        current_row += 1
        
        # Headers de la tabla de recursos (7 columnas)
        headers = ["CÓDIGO RECURSO", "DESCRIPCIÓN", "UND", "CANT.", "DESP.%", "PRECIO UNIT", "VALOR PARCIAL"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=current_row, column=col, value=header)
            ws.cell(row=current_row, column=col).font = small_font
            ws.cell(row=current_row, column=col).fill = light_fill
            ws.cell(row=current_row, column=col).alignment = center_align
            ws.cell(row=current_row, column=col).border = thin_border
        current_row += 1
        
        # Obtener recursos del análisis
        recursos = self._get_recursos_for_analisis(analisis_data['codigo'])
        
        # Debug: si no hay recursos, mostrar mensaje
        if not recursos:
            ws.cell(row=current_row, column=1, value=f"(No se encontraron recursos para el análisis {analisis_data['codigo']})")
            ws.cell(row=current_row, column=1).font = small_font
            for col in range(1, 7):
                ws.cell(row=current_row, column=col).border = thin_border
            current_row += 1
        
        # Subtotales por categoría
        subtotales = {'MATERIALES': 0, 'MANO DE OBRA': 0, 'EQUIPO': 0, 'OTROS': 0}
        current_category = None
        
        for recurso in recursos:
            # Detectar cambio de categoría
            if recurso['categoria'] != current_category:
                if current_category:  # Agregar subtotal de categoría anterior
                    ws.merge_cells(f'A{current_row}:F{current_row}')
                    ws.cell(row=current_row, column=1, value=f"SUBTOTAL {current_category}:")
                    ws.cell(row=current_row, column=1).font = small_font
                    ws.cell(row=current_row, column=1).alignment = right_align
                    ws.cell(row=current_row, column=7, value=subtotales[current_category])
                    ws.cell(row=current_row, column=7).font = small_font
                    ws.cell(row=current_row, column=7).number_format = '#,##0.00'
                    ws.cell(row=current_row, column=7).alignment = right_align
                    for col in range(1, 8):
                        ws.cell(row=current_row, column=col).border = thin_border
                        ws.cell(row=current_row, column=col).fill = light_fill
                    current_row += 1
                
                # Header de nueva categoría
                current_category = recurso['categoria']
                ws.cell(row=current_row, column=1, value=f"--- {current_category} ---")
                ws.cell(row=current_row, column=1).font = small_font
                ws.cell(row=current_row, column=1).alignment = left_align
                for col in range(1, 8):
                    ws.cell(row=current_row, column=col).border = thin_border
                current_row += 1
            
            # Datos del recurso
            valor_parcial = recurso['cantidad'] * (1 + recurso['desperdicio']/100) * recurso['precio_unitario']
            subtotales[current_category] += valor_parcial
            
            ws.cell(row=current_row, column=1, value=recurso['codigo'])
            ws.cell(row=current_row, column=2, value=recurso['descripcion'])
            ws.cell(row=current_row, column=3, value=recurso['unidad'])
            ws.cell(row=current_row, column=4, value=recurso['cantidad'])
            ws.cell(row=current_row, column=5, value=f"{recurso['desperdicio']:.1f}%")
            ws.cell(row=current_row, column=6, value=recurso['precio_unitario'])
            ws.cell(row=current_row, column=7, value=valor_parcial)
            
            # Aplicar formato a todas las 7 columnas
            for col in range(1, 8):
                ws.cell(row=current_row, column=col).font = small_font
                ws.cell(row=current_row, column=col).border = thin_border
                if col in [4, 6, 7]:  # Cantidad, Precio Unit, Valor Parcial
                    ws.cell(row=current_row, column=col).number_format = '#,##0.00'
                    ws.cell(row=current_row, column=col).alignment = right_align
                elif col == 5:  # Desperdicio %
                    ws.cell(row=current_row, column=col).alignment = center_align
                else:  # Código, Descripción, Unidad
                    ws.cell(row=current_row, column=col).alignment = left_align
            
            current_row += 1
        
        # Subtotal de la última categoría
        if current_category:
            ws.merge_cells(f'A{current_row}:F{current_row}')
            ws.cell(row=current_row, column=1, value=f"SUBTOTAL {current_category}:")
            ws.cell(row=current_row, column=1).font = small_font
            ws.cell(row=current_row, column=1).alignment = right_align
            ws.cell(row=current_row, column=7, value=subtotales[current_category])
            ws.cell(row=current_row, column=7).font = small_font
            ws.cell(row=current_row, column=7).number_format = '#,##0.00'
            ws.cell(row=current_row, column=7).alignment = right_align
            for col in range(1, 8):
                ws.cell(row=current_row, column=col).border = thin_border
                ws.cell(row=current_row, column=col).fill = light_fill
            current_row += 1
        
        # Costo directo total
        costo_directo = sum(subtotales.values())
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws.cell(row=current_row, column=1, value="COSTO DIRECTO:")
        ws.cell(row=current_row, column=1).font = header_font
        ws.cell(row=current_row, column=1).alignment = right_align
        ws.cell(row=current_row, column=7, value=costo_directo)
        ws.cell(row=current_row, column=7).font = header_font
        ws.cell(row=current_row, column=7).number_format = '#,##0.00'
        ws.cell(row=current_row, column=7).alignment = right_align
        for col in range(1, 8):
            ws.cell(row=current_row, column=col).border = thin_border
            ws.cell(row=current_row, column=col).fill = header_fill
        current_row += 1
        
        # Costos indirectos
        current_row += 1
        ws.merge_cells(f'A{current_row}:G{current_row}')
        ws.cell(row=current_row, column=1, value="COSTOS INDIRECTOS")
        ws.cell(row=current_row, column=1).font = header_font
        ws.cell(row=current_row, column=1).alignment = center_align
        for col in range(1, 8):
            ws.cell(row=current_row, column=col).border = thin_border
            ws.cell(row=current_row, column=col).fill = header_fill
        current_row += 1
        
        # Aplicar porcentajes AIU
        aiu_items = [
            ("ADMINISTRACIÓN", aiu_percentages['admin_pct']),
            ("IMPREVISTOS", aiu_percentages['imprev_pct']),
            ("UTILIDAD", aiu_percentages['util_pct']),
            ("IVA SOBRE LA UTILIDAD", aiu_percentages['iva_pct']),
            ("TOTAL COSTOS INDIRECTOS", aiu_percentages['total_indirectos_pct'])
        ]
        
        for concepto, porcentaje in aiu_items:
            valor = costo_directo * (porcentaje / 100)
            
            ws.merge_cells(f'A{current_row}:E{current_row}')
            ws.cell(row=current_row, column=1, value=concepto)
            ws.cell(row=current_row, column=1).font = normal_font
            ws.cell(row=current_row, column=1).alignment = right_align
            
            ws.cell(row=current_row, column=6, value=f"{porcentaje:.2f}%")
            ws.cell(row=current_row, column=6).font = normal_font
            ws.cell(row=current_row, column=6).alignment = center_align
            
            ws.cell(row=current_row, column=7, value=valor)
            ws.cell(row=current_row, column=7).font = normal_font
            ws.cell(row=current_row, column=7).number_format = '#,##0.00'
            ws.cell(row=current_row, column=7).alignment = right_align
            
            for col in range(1, 8):
                ws.cell(row=current_row, column=col).border = thin_border
            
            current_row += 1
        
        # Valor total del ítem
        valor_total_item = costo_directo * (1 + aiu_percentages['total_indirectos_pct'] / 100)
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws.cell(row=current_row, column=1, value="VALOR TOTAL ITEM:")
        ws.cell(row=current_row, column=1).font = header_font
        ws.cell(row=current_row, column=1).alignment = right_align
        ws.cell(row=current_row, column=7, value=valor_total_item)
        ws.cell(row=current_row, column=7).font = header_font
        ws.cell(row=current_row, column=7).number_format = '#,##0.00'
        ws.cell(row=current_row, column=7).alignment = right_align
        for col in range(1, 8):
            ws.cell(row=current_row, column=col).border = thin_border
            ws.cell(row=current_row, column=col).fill = header_fill
        
        return current_row + 1
    
    def _get_recursos_for_analisis(self, codigo_analisis):
        """Obtiene los recursos desde la vista de análisis del presupuesto"""
        try:
            # Buscar en la ventana principal los controladores de análisis
            from PyQt6.QtWidgets import QApplication
            app = QApplication.instance()
            if app:
                for widget in app.topLevelWidgets():
                    if hasattr(widget, '_analisis_presupuesto_ctrls'):
                        # Buscar el controlador específico para este análisis
                        for ctrl in widget._analisis_presupuesto_ctrls:
                            if hasattr(ctrl, 'codigo_analisis') and ctrl.codigo_analisis == codigo_analisis:
                                return self._extract_recursos_from_controller(ctrl)
            
            # Si no encuentra el controlador, usar datos de ejemplo
            return self._get_ejemplo_recursos(codigo_analisis)
            
        except Exception as e:
            print(f"Error obteniendo recursos para {codigo_analisis}: {e}")
            return self._get_ejemplo_recursos(codigo_analisis)
    
    def _extract_recursos_from_controller(self, controller):
        """Extrae recursos del controlador de RecursosPorAnalisis"""
        recursos = []
        try:
            if hasattr(controller, 'view') and hasattr(controller.view, 'model'):
                model = controller.view.model
                current_category = 'OTROS'
                
                for row in range(model.rowCount()):
                    # Obtener datos de cada columna
                    codigo_item = model.item(row, 0)
                    desc_item = model.item(row, 1)
                    unidad_item = model.item(row, 2)
                    cantidad_item = model.item(row, 3)
                    desp_item = model.item(row, 4)
                    precio_item = model.item(row, 5)
                    
                    if not codigo_item:
                        continue
                    
                    codigo = codigo_item.text()
                    
                    # Detectar headers de categoría
                    if codigo.startswith('===') or codigo.startswith('===='):
                        if 'MANO DE OBRA' in codigo:
                            current_category = 'MANO DE OBRA'
                        elif 'EQUIPO' in codigo:
                            current_category = 'EQUIPO'
                        elif 'MATERIALES' in codigo:
                            current_category = 'MATERIALES'
                        else:
                            current_category = 'OTROS'
                        continue
                    
                    # Extraer datos del recurso
                    descripcion = desc_item.text() if desc_item else ''
                    unidad = unidad_item.text() if unidad_item else ''
                    cantidad_str = cantidad_item.text() if cantidad_item else '0'
                    desp_str = desp_item.text() if desp_item else '0'
                    precio_str = precio_item.text() if precio_item else '$0'
                    
                    # Limpiar y convertir valores
                    try:
                        cantidad = float(cantidad_str.replace(',', ''))
                    except:
                        cantidad = 0.0
                    
                    try:
                        desperdicio = float(desp_str.replace('%', '').replace(',', ''))
                    except:
                        desperdicio = 0.0
                    
                    try:
                        precio = float(precio_str.replace('$', '').replace(',', ''))
                    except:
                        precio = 0.0
                    
                    if codigo and descripcion:  # Solo agregar si tiene datos válidos
                        recursos.append({
                            'codigo': codigo,
                            'descripcion': descripcion,
                            'unidad': unidad,
                            'cantidad': cantidad,
                            'desperdicio': desperdicio,
                            'precio_unitario': precio,
                            'categoria': current_category
                        })
                
        except Exception as e:
            print(f"Error extrayendo recursos del controlador: {e}")
        
        return recursos
    
    def _create_insumos_sheet(self, workbook):
        """Crea la hoja de INSUMOS consolidando todos los recursos por categoría"""
        ws = workbook.create_sheet("INSUMOS")
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 50  # DESCRIPCIÓN
        ws.column_dimensions['B'].width = 8   # UND
        ws.column_dimensions['C'].width = 12  # CANT.
        ws.column_dimensions['D'].width = 15  # VR. UNIT.
        ws.column_dimensions['E'].width = 15  # VR. TOTAL
        
        # Estilos
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        header_font = Font(bold=True, size=12)
        normal_font = Font(size=10)
        small_font = Font(size=9)
        
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        blue_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Header del documento
        ws.merge_cells('A1:C1')
        ws['A1'] = "NOMBRE DE LA EMPRESA"
        ws['A1'].font = header_font
        ws['A1'].alignment = center_align
        ws['A1'].border = thin_border
        
        ws.merge_cells('D1:E1')
        ws['D1'] = "LOGO DE LA EMPRESA"
        ws['D1'].font = header_font
        ws['D1'].alignment = center_align
        ws['D1'].border = thin_border
        
        ws.row_dimensions[1].height = 25
        
        # Título principal
        current_row = 3
        ws.merge_cells(f'A{current_row}:E{current_row}')
        ws.cell(row=current_row, column=1, value="INSUMOS")
        ws.cell(row=current_row, column=1).font = header_font
        ws.cell(row=current_row, column=1).alignment = center_align
        ws.cell(row=current_row, column=1).fill = blue_fill
        for col in range(1, 6):
            ws.cell(row=current_row, column=col).border = thin_border
        
        # Información del proyecto
        current_row += 2
        ws.cell(row=current_row, column=1, value="Obra:")
        ws.cell(row=current_row, column=2, value="ESCRIBA AQUÍ EL NOMBRE DE LA OBRA")
        ws.cell(row=current_row, column=4, value="FECHA:")
        ws.cell(row=current_row, column=5, value="19-sept-25")
        ws.cell(row=current_row+1, column=4, value="QUIEN ELABORÓ:")
        
        current_row += 3
        
        # Consolidar recursos de todos los análisis
        consolidated_resources = self._consolidate_all_resources()
        
        # Crear tabla por cada categoría
        for categoria in ['MATERIALES', 'MANO DE OBRA', 'EQUIPO']:
            if categoria in consolidated_resources:
                current_row = self._create_categoria_insumos_table(
                    ws, categoria, consolidated_resources[categoria], current_row,
                    header_font, normal_font, small_font, header_fill, light_fill, blue_fill,
                    center_align, left_align, right_align, thin_border
                )
        
        # Total general
        total_general = sum(
            sum(recurso['valor_total'] for recurso in recursos.values())
            for recursos in consolidated_resources.values()
        )
        
        current_row += 1
        ws.merge_cells(f'A{current_row}:D{current_row}')
        ws.cell(row=current_row, column=1, value="TOTAL GENERAL:")
        ws.cell(row=current_row, column=1).font = header_font
        ws.cell(row=current_row, column=1).alignment = right_align
        ws.cell(row=current_row, column=5, value=total_general)
        ws.cell(row=current_row, column=5).font = header_font
        ws.cell(row=current_row, column=5).number_format = '#,##0.00'
        ws.cell(row=current_row, column=5).alignment = right_align
        
        for col in range(1, 6):
            ws.cell(row=current_row, column=col).border = thin_border
            ws.cell(row=current_row, column=col).fill = header_fill
    
    def _consolidate_all_resources(self):
        """Consolida todos los recursos de todos los análisis, sumando cantidades duplicadas"""
        consolidated = {
            'MATERIALES': {},
            'MANO DE OBRA': {},
            'EQUIPO': {}
        }
        
        # Obtener análisis únicos del presupuesto
        analisis_list = self._get_unique_analisis_from_presupuesto()
        
        for analisis_data in analisis_list:
            recursos = self._get_recursos_for_analisis(analisis_data['codigo'])
            # IMPORTANTE: Multiplicar por la cantidad del análisis en el presupuesto
            cantidad_analisis = analisis_data.get('cantidad', 1.0)
            
            for recurso in recursos:
                categoria = recurso['categoria']
                if categoria not in consolidated:
                    categoria = 'MATERIALES'  # Default
                
                key = f"{recurso['codigo']}_{recurso['descripcion']}"
                
                # Cantidad real = cantidad del recurso × cantidad del análisis en presupuesto
                cantidad_real = recurso['cantidad'] * cantidad_analisis
                
                if key in consolidated[categoria]:
                    # Sumar cantidad si ya existe
                    consolidated[categoria][key]['cantidad'] += cantidad_real
                    # Recalcular valor total
                    consolidated[categoria][key]['valor_total'] = (
                        consolidated[categoria][key]['cantidad'] * 
                        (1 + consolidated[categoria][key]['desperdicio']/100) * 
                        consolidated[categoria][key]['precio_unitario']
                    )
                else:
                    # Nuevo recurso
                    consolidated[categoria][key] = {
                        'codigo': recurso['codigo'],
                        'descripcion': recurso['descripcion'],
                        'unidad': recurso['unidad'],
                        'cantidad': cantidad_real,  # Ya multiplicada por cantidad del análisis
                        'desperdicio': recurso['desperdicio'],
                        'precio_unitario': recurso['precio_unitario'],
                        'valor_total': cantidad_real * (1 + recurso['desperdicio']/100) * recurso['precio_unitario']
                    }
        
        return consolidated
    
    def _create_categoria_insumos_table(self, ws, categoria, recursos, start_row,
                                       header_font, normal_font, small_font,
                                       header_fill, light_fill, blue_fill, center_align, 
                                       left_align, right_align, thin_border):
        """Crea una tabla para una categoría específica de insumos"""
        current_row = start_row
        
        # Header de categoría
        ws.merge_cells(f'A{current_row}:E{current_row}')
        ws.cell(row=current_row, column=1, value=categoria)
        ws.cell(row=current_row, column=1).font = header_font
        ws.cell(row=current_row, column=1).alignment = center_align
        ws.cell(row=current_row, column=1).fill = blue_fill
        for col in range(1, 6):
            ws.cell(row=current_row, column=col).border = thin_border
        current_row += 1
        
        # Headers de la tabla
        headers = ["DESCRIPCIÓN", "UND", "CANT.", "VR. UNIT.", "VR.TOTAL"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=current_row, column=col, value=header)
            ws.cell(row=current_row, column=col).font = small_font
            ws.cell(row=current_row, column=col).fill = light_fill
            ws.cell(row=current_row, column=col).alignment = center_align
            ws.cell(row=current_row, column=col).border = thin_border
        current_row += 1
        
        # Datos de recursos
        subtotal = 0
        for recurso in recursos.values():
            ws.cell(row=current_row, column=1, value=recurso['descripcion'])
            ws.cell(row=current_row, column=2, value=recurso['unidad'])
            ws.cell(row=current_row, column=3, value=recurso['cantidad'])
            ws.cell(row=current_row, column=4, value=recurso['precio_unitario'])
            ws.cell(row=current_row, column=5, value=recurso['valor_total'])
            
            subtotal += recurso['valor_total']
            
            # Aplicar formato
            for col in range(1, 6):
                ws.cell(row=current_row, column=col).font = small_font
                ws.cell(row=current_row, column=col).border = thin_border
                if col == 1:  # Descripción
                    ws.cell(row=current_row, column=col).alignment = left_align
                elif col in [3, 4, 5]:  # Cantidades y valores
                    ws.cell(row=current_row, column=col).number_format = '#,##0.00'
                    ws.cell(row=current_row, column=col).alignment = right_align
                else:  # Unidad
                    ws.cell(row=current_row, column=col).alignment = center_align
            
            current_row += 1
        
        # Subtotal
        ws.merge_cells(f'A{current_row}:D{current_row}')
        ws.cell(row=current_row, column=1, value=f"SUBTOTAL {categoria}:")
        ws.cell(row=current_row, column=1).font = normal_font
        ws.cell(row=current_row, column=1).alignment = right_align
        ws.cell(row=current_row, column=5, value=subtotal)
        ws.cell(row=current_row, column=5).font = normal_font
        ws.cell(row=current_row, column=5).number_format = '#,##0.00'
        ws.cell(row=current_row, column=5).alignment = right_align
        
        for col in range(1, 6):
            ws.cell(row=current_row, column=col).border = thin_border
            ws.cell(row=current_row, column=col).fill = light_fill
        
        return current_row + 2  # Espacio para la siguiente categoría
    
    def _get_ejemplo_recursos(self, codigo_analisis):
        """Genera recursos de ejemplo si no se pueden obtener de la BD"""
        # Recursos de ejemplo basados en el código del análisis
        if "MOTOR" in codigo_analisis.upper():
            return [
                {
                    'codigo': 'MOI501',
                    'descripcion': 'MANO OBRA HIDROSANIT.1 AYUDANTE-1 OFI',
                    'unidad': 'HC',
                    'cantidad': 5.5,
                    'desperdicio': 0.0,
                    'precio_unitario': 25391.00,
                    'categoria': 'MANO DE OBRA'
                },
                {
                    'codigo': 'MQ0301',
                    'descripcion': 'HERRAMIENTA MENOR',
                    'unidad': 'GLB',
                    'cantidad': 1.9,
                    'desperdicio': 0.0,
                    'precio_unitario': 1600.00,
                    'categoria': 'EQUIPO'
                },
                {
                    'codigo': '004837',
                    'descripcion': 'MOTOR EL.100HP 3600RPM BOM',
                    'unidad': 'UND',
                    'cantidad': 1.0,
                    'desperdicio': 0.0,
                    'precio_unitario': 11024350.00,
                    'categoria': 'MATERIALES'
                }
            ]
        elif "BOMBA" in codigo_analisis.upper():
            return [
                {
                    'codigo': 'MOI501',
                    'descripcion': 'MANO OBRA ELECTRICAS1 AYUDANTE-1 OFI',
                    'unidad': 'HC',
                    'cantidad': 2.5,
                    'desperdicio': 0.0,
                    'precio_unitario': 33455.00,
                    'categoria': 'MANO DE OBRA'
                },
                {
                    'codigo': '004857',
                    'descripcion': 'BOM.SUM. 15 HP LAPIC 6"',
                    'unidad': 'UND',
                    'cantidad': 1.0,
                    'desperdicio': 0.0,
                    'precio_unitario': 11008400.00,
                    'categoria': 'MATERIALES'
                }
            ]
        elif "TUB" in codigo_analisis.upper():
            return [
                {
                    'codigo': 'ALU501',
                    'descripcion': 'TUBERIA PVC 4" NOVAFORT',
                    'unidad': 'ML',
                    'cantidad': 1.0,
                    'desperdicio': 5.0,
                    'precio_unitario': 142682.00,
                    'categoria': 'MATERIALES'
                },
                {
                    'codigo': 'MOI301',
                    'descripcion': 'MANO OBRA PLOMERIA-1 OFI',
                    'unidad': 'HC',
                    'cantidad': 0.5,
                    'desperdicio': 0.0,
                    'precio_unitario': 28500.00,
                    'categoria': 'MANO DE OBRA'
                }
            ]
        else:
            return [
                {
                    'codigo': 'GENERICO',
                    'descripcion': f'Material para {codigo_analisis}',
                    'unidad': 'UND',
                    'cantidad': 1.0,
                    'desperdicio': 0.0,
                    'precio_unitario': 100000.00,
                    'categoria': 'MATERIALES'
                }
            ]
